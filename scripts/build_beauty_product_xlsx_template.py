# -*- coding: utf-8 -*-
"""Generates static/downloads/beauty_product_template.xlsx for beauty center inventory."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "static" / "downloads" / "beauty_product_template.xlsx"

HEADERS = [
    ("اسم المادة أو المنتج (إلزامي)", "name", "نص — مثال: سيروم فيتامين C"),
    ("مادة للجلسات فقط ولا تباع للزبون", "sessions_only", "نعم / لا — نعم = لا تُباع في الواجهة"),
    ("الكمية الحالية", "opening_stock", "رقم صحيح"),
    ("تنبيه الكمية (حد التنبيه)", "low_stock_threshold", "مثال: 5"),
    ("كلفة المنتج أو المادة (د.ع)", "buy_price", "إلزامي — رقم"),
    ("سعر البيع للزبون (د.ع)", "sale_price", "اختياري للمواد الجلسات فقط"),
    ("نوع المنتج", "beauty_product_type", "انظر ورقة «قيم مسموحة»"),
    ("نوع البشرة المناسب", "skin_type", "انظر ورقة «قيم مسموحة»"),
    ("منطقة الاستخدام", "usage_type", "انظر ورقة «قيم مسموحة»"),
    ("الشركة / الماركة", "brand", "نص"),
    ("رقم التشغيلة", "batch_number", "نص"),
    ("يتطلب Patch Test", "requires_patch_test", "لا / نعم"),
    ("مدة الصلاحية بعد الفتح (أيام)", "shelf_life_after_opening_days", "مثال: 90"),
    ("تاريخ الفتح", "opened_date", "YYYY-MM-DD"),
    ("تاريخ الانتهاء", "expiry_date", "YYYY-MM-DD"),
    ("طريقة الاستخدام المختصرة", "usage_instructions", "نص متعدد الأسطر"),
    ("ملاحظات تحذيرية", "warning_notes", "نص متعدد الأسطر"),
    ("وصف المادة / المنتج", "description", "اختياري"),
    ("رابط صورة بديل (اختياري)", "external_image_url", "https://..."),
]

ALLOWED = [
    ("beauty_product_type", "session_material", "مادة جلسات"),
    ("beauty_product_type", "retail_product", "منتج بيع للزبون"),
    ("beauty_product_type", "tool", "أداة"),
    ("beauty_product_type", "consumable", "مستهلكات"),
    ("skin_type", "", "غير محدد"),
    ("skin_type", "all", "كل الأنواع"),
    ("skin_type", "oily", "دهنية"),
    ("skin_type", "dry", "جافة"),
    ("skin_type", "combination", "مختلطة"),
    ("skin_type", "sensitive", "حساسة"),
    ("usage_type", "", "غير محدد"),
    ("usage_type", "face", "وجه"),
    ("usage_type", "body", "جسم"),
    ("usage_type", "hair", "شعر"),
    ("usage_type", "laser", "ليزر"),
    ("usage_type", "general_care", "عناية عامة"),
]


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "قالب الإدخال"

    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="64748b")

    for col, (ar_label, key, hint) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=ar_label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=col, value=f"{key}\n{hint}")
        ws.cell(row=2, column=col).font = Font(size=9, color="64748b")
        ws.cell(row=2, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # صف مثال (اختياري للمستخدم)
    example = [
        "سيروم فيتامين C",
        "لا",
        12,
        5,
        45000,
        65000,
        "retail_product",
        "combination",
        "face",
        "BrandX",
        "BATCH-2026-01",
        "لا",
        90,
        "2026-01-15",
        "2027-06-01",
        "يُستخدم بعد تنظيف البشرة.",
        "يُمنع على البشرة المتهيجة.",
        "وصف قصير للمادة.",
        "",
    ]
    for col, val in enumerate(example, start=1):
        c = ws.cell(row=3, column=col, value=val)
        c.font = Font(italic=True, color="475569")
        c.border = Border(bottom=thin)

    ws.freeze_panes = "A4"
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(28, 14 + len(HEADERS[col - 1][0]) // 3)

    # ورقة القيم
    w2 = wb.create_sheet("قيم مسموحة")
    w2.append(["الحقل", "القيمة في النظام (بالإنجليزية)", "معنى عربي"])
    for row in ALLOWED:
        w2.append(list(row))
    for cell in w2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="e2e8f0")

    w3 = wb.create_sheet("ملاحظات")
    w3.merge_cells("A1:B1")
    w3["A1"] = "تعليمات استخدام القالب — مركز التجميل"
    w3["A1"].font = Font(bold=True, size=14)
    notes = [
        ("الصف 1", "عناوين الأعمدة كما في شاشة إضافة المنتج."),
        ("الصف 2", "المعرّف الداخلي للحقل + تلميح مختصر."),
        ("الصف 3", "صف مثال يمكن حذفه أو استبداله."),
        ("الصف 4 فما فوق", "أدخل المواد — صف واحد لكل منتج."),
        ("التواريخ", "استخدم الصيغة YYYY-MM-DD حتى يُستورد لاحقاً بسهولة."),
        ("نعم/لا", "للحقول المنطقية: نعم أو لا (أو 1/0 إن رغبت في الاستيراد البرمجي)."),
        ("الصورة", "لا يُرفع ملف عبر Excel؛ استخدم عمود رابط الصورة أو أضف الصورة يدوياً من النظام."),
        ("استيراد تلقائي", "هذا القالب للتعبئة اليدوية أو لدمجها مع أداة استيراد مستقبلية؛ الصق القيم كما في ورقة «قيم مسموحة»."),
    ]
    for i, (k, v) in enumerate(notes, start=3):
        w3.cell(row=i, column=1, value=k).font = Font(bold=True)
        w3.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True)
    w3.column_dimensions["A"].width = 22
    w3.column_dimensions["B"].width = 72

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
