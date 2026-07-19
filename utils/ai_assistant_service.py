"""AI assistant orchestration, Excel audit parsing, and approved action execution."""
from __future__ import annotations

import json
import os
import re
import uuid
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from flask import current_app, has_request_context, session
from sqlalchemy import func, inspect, or_, text
from werkzeug.datastructures import FileStorage
from werkzeug.utils import secure_filename

from extensions import db
from models.ai_assistant_control import (
    AIActionItem,
    AIActionPlan,
    AIAuditRun,
    AIChatMessage,
    AIChatSession,
    AIScheduledAudit,
    AIToolCallLog,
    AIUploadedFile,
)
from models.branch import Branch, BranchStock, StockTransfer, StockTransferLine
from models.employee import Employee
from models.fixed_asset import FixedAsset
from models.fixed_asset_category import FixedAssetCategory
from models.invoice import Invoice
from models.message import Message
from models.order_item import OrderItem
from models.product import Product
from models.role import Permission
from models.shipping import ShippingCompany
from models.shipping_payment import ShippingPayment
from models.supplier import Supplier
from models.system_alert import SystemAlert
from models.system_analytics import SystemAnalytics
from utils.audit_accounting_integrity import _ensure_audit_schema, audit_accounting_integrity
from utils.branch_migration import ensure_branch_schema, get_default_branch
from utils.cash_calculations import _effective_paid_amount
from utils.delivery_expense_service import sync_delivery_expense_for_invoice
from utils.order_lifecycle import OrderLifecycleError, process_order_cancel, process_order_return
from utils.order_stock_policy import ensure_stock_for_transition
from utils.branch_stock_service import (
    BranchStockError,
    adjust_branch_stock,
    get_branch_stock,
    sync_product_total,
    transfer_deduct,
    transfer_receive,
)
from utils.inventory_movements import get_product_inventory_movements
from utils.order_item_costs import exclude_delivery_fee_items
from utils.activity_logger import log_activity
from utils.ai_assistant_tools import (
    execute_tool,
    get_tool_definitions,
    now_local as business_now_local,
)
from utils.payment_ledger import append_payment_ledger_delta
from utils.permission_checks import employee_can
from utils.product_schema_guard import ensure_product_schema
from utils.supplier_accounting_repair import audit_and_repair_supplier_ledgers
from utils.fixed_assets_service import (
    build_asset_from_form,
    post_asset_acquisition,
    seed_default_categories,
)
from utils.treasury_calculations import calculate_treasury_balance
from utils.treasury_helpers import get_default_cash_account


AI_PERMISSION_DEFS = [
    ("use_ai_assistant", "استخدام مساعد Finora الذكي"),
    ("approve_ai_actions", "الموافقة على خطط المساعد الذكي وتنفيذها"),
    ("manage_ai_schedules", "إدارة دوريات تحليل المساعد الذكي"),
    ("view_ai_audit_logs", "رؤية سجل أدوات ومراجعات المساعد الذكي"),
]

RESERVED_ORDER_STATUSES = ("تم الطلب", "جاري الشحن")
ALLOWED_UPLOAD_EXTENSIONS = {".xlsx", ".xlsm"}


def ensure_ai_permissions() -> None:
    created = False
    for name, description in AI_PERMISSION_DEFS:
        if not Permission.query.filter_by(name=name).first():
            db.session.add(Permission(name=name, description=description, created_at=datetime.utcnow()))
            created = True
    if created:
        db.session.commit()


def ensure_ai_assistant_schema() -> None:
    """Create assistant tables and seed permissions for deployments without migrations."""
    from models import ai_assistant_control  # noqa: F401

    bind = db.session.get_bind()
    db.Model.metadata.create_all(
        bind=bind,
        tables=[
            AIChatSession.__table__,
            AIChatMessage.__table__,
            AIUploadedFile.__table__,
            AIActionPlan.__table__,
            AIActionItem.__table__,
            AIScheduledAudit.__table__,
            AIAuditRun.__table__,
            AIToolCallLog.__table__,
        ],
    )
    _ensure_message_schema_for_ai_alerts()
    ensure_ai_permissions()


def _ensure_message_schema_for_ai_alerts() -> None:
    """AI critical alerts write internal messages; keep legacy tenant DBs compatible."""
    try:
        bind = db.session.get_bind()
        inspector = inspect(bind)
        if "message" not in inspector.get_table_names():
            return
        columns = {col["name"] for col in inspector.get_columns("message")}
        additions = {
            "file_type": "ALTER TABLE message ADD COLUMN file_type VARCHAR(50)",
            "file_path": "ALTER TABLE message ADD COLUMN file_path VARCHAR(500)",
            "file_name": "ALTER TABLE message ADD COLUMN file_name VARCHAR(255)",
            "is_edited": "ALTER TABLE message ADD COLUMN is_edited BOOLEAN DEFAULT 0",
            "reply_to_id": "ALTER TABLE message ADD COLUMN reply_to_id INTEGER",
        }
        changed = False
        for column, stmt in additions.items():
            if column not in columns:
                db.session.execute(text(stmt))
                changed = True
        if changed:
            db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("failed to ensure message schema for AI alerts")


def _json_dumps(data: Any) -> str:
    return json.dumps(data or {}, ensure_ascii=False)


def _current_employee_id() -> int | None:
    if has_request_context():
        return session.get("user_id")
    return None


def _serialize_plan(plan: AIActionPlan | None) -> dict | None:
    return plan.to_dict() if plan else None


def _log_tool(
    tool_name: str,
    *,
    session_id: int | None = None,
    plan_id: int | None = None,
    employee_id: int | None = None,
    mode: str = "read",
    input_data: dict | None = None,
    output_data: dict | None = None,
    status: str = "success",
    error: str | None = None,
) -> None:
    try:
        row = AIToolCallLog(
            session_id=session_id,
            plan_id=plan_id,
            employee_id=employee_id,
            tool_name=tool_name[:120],
            mode=mode,
            status=status,
            error_message=error,
        )
        row.set_input(input_data or {})
        row.set_output(output_data or {})
        db.session.add(row)
    except Exception:
        current_app.logger.exception("failed to log AI tool call")


def get_or_create_chat_session(employee_id: int | None, session_id: int | None = None) -> AIChatSession:
    if session_id:
        existing = AIChatSession.query.get(session_id)
        if existing and (existing.employee_id == employee_id or employee_id is None):
            return existing
    chat = AIChatSession(employee_id=employee_id, title="محادثة مساعد Finora")
    chat.set_context({"source": "assistant_chat"})
    db.session.add(chat)
    db.session.flush()
    return chat


def add_chat_message(chat: AIChatSession, role: str, content: str, metadata: dict | None = None) -> AIChatMessage:
    msg = AIChatMessage(session_id=chat.id, role=role, content=content or "")
    msg.set_metadata(metadata or {})
    db.session.add(msg)
    chat.updated_at = datetime.utcnow()
    db.session.flush()
    return msg


def _money(value) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _assistant_read_scope(employee_id: int | None = None) -> dict:
    employee = Employee.query.get(employee_id) if employee_id else None
    if not employee:
        return {
            "full": True,
            "orders": True,
            "reports": True,
            "financial": True,
            "inventory": True,
            "suppliers": True,
            "shipping": True,
            "assets": True,
            "order_manage": True,
        }
    is_admin = employee.role == "admin"
    return {
        "full": is_admin,
        "orders": is_admin or employee_can(employee, "view_orders"),
        "reports": is_admin or employee_can(employee, "view_reports"),
        "financial": is_admin or employee_can(employee, "view_financial") or employee_can(employee, "view_accounts"),
        "inventory": is_admin or employee_can(employee, "manage_inventory"),
        "suppliers": is_admin or employee_can(employee, "manage_suppliers"),
        "shipping": is_admin or employee_can(employee, "view_shipping") or employee_can(employee, "manage_shipping"),
        "assets": is_admin or employee_can(employee, "view_fixed_assets") or employee_can(employee, "manage_fixed_assets"),
        "order_manage": is_admin or employee_can(employee, "manage_orders"),
    }


def _restricted(label: str = "غير متاح حسب صلاحياتك") -> dict:
    return {"restricted": True, "message": label}


def collect_system_snapshot(employee_id: int | None = None) -> dict:
    _ensure_audit_schema()
    ensure_product_schema()
    ensure_branch_schema()
    scope = _assistant_read_scope(employee_id)
    try:
        from utils.accounting_calculations import (
            calculate_shipping_due,
            calculate_shipping_opening_balance,
            calculate_supplier_debts,
        )

        supplier_debts = _money(calculate_supplier_debts())
        shipping_receivables = _money(calculate_shipping_due())
        shipping_opening_balance = _money(calculate_shipping_opening_balance())
    except Exception:
        supplier_debts = 0
        shipping_receivables = 0
        shipping_opening_balance = 0
    total_products = Product.query.count() if scope["inventory"] else None
    active_products = Product.query.filter_by(active=True).count() if scope["inventory"] else None
    branches = Branch.query.order_by(Branch.name.asc()).all() if scope["inventory"] else []
    stock_value = (
        db.session.query(func.coalesce(func.sum(Product.quantity * Product.buy_price), 0)).scalar() or 0
        if scope["inventory"] and (scope["financial"] or scope["reports"])
        else None
    )
    orders_by_status = (
        dict(db.session.query(Invoice.status, func.count(Invoice.id)).group_by(Invoice.status).all())
        if scope["orders"]
        else {}
    )
    recent_sales = (
        db.session.query(func.coalesce(func.sum(Invoice.total), 0))
        .filter(
            Invoice.created_at >= datetime.utcnow() - timedelta(days=30),
            Invoice.status.notin_(["ملغي", "مرتجع", "راجع", "راجعة"]),
        )
        .scalar()
        or 0
        if scope["reports"] or scope["financial"]
        else None
    )
    negative_margin_count = (
        db.session.query(OrderItem.id)
        .join(Invoice, Invoice.id == OrderItem.invoice_id)
        .filter(
            Invoice.status.notin_(["ملغي", "مرتجع", "راجع", "راجعة"]),
            OrderItem.cost > OrderItem.price,
        )
        .count()
        if scope["financial"] or scope["reports"]
        else None
    )
    reserved = _reserved_stock_map() if scope["inventory"] else {}
    financial_data = {}
    if scope["financial"] or scope["suppliers"]:
        financial_data["supplier_debts"] = supplier_debts
    if scope["financial"] or scope["shipping"]:
        financial_data["shipping_receivables"] = shipping_receivables
        financial_data["shipping_opening_balance_receivable"] = shipping_opening_balance
        financial_data["shipping_note"] = "مستحقات شركات النقل ذمم مدينة لصالح الشركة، وليست ديناً على الشركة."
    if scope["financial"]:
        try:
            from utils.cash_calculations import get_cash_summary

            cash_summary = get_cash_summary()
            financial_data["cash"] = {
                "balance": _money(cash_summary.get("current_balance")),
                "total_in": _money(cash_summary.get("total_in")),
                "total_out": _money(cash_summary.get("total_out")),
                "movements_count": int(cash_summary.get("movements_count") or 0),
                "audit_rule": "أي حركة صندوق يدوية بلا سبب/ملاحظة تعتبر مشبوهة وتحتاج مراجعة قبل اعتمادها.",
            }
        except Exception as exc:
            financial_data["cash"] = {"error": str(exc)}
    if not financial_data:
        financial_data = _restricted()
    return {
        "generated_at": datetime.utcnow().isoformat(),
        "permission_scope": scope,
        "products": (
            {"total": total_products, "active": active_products, "stock_value": _money(stock_value)}
            if scope["inventory"]
            else _restricted()
        ),
        "branches": [{"id": b.id, "name": b.name, "code": b.code} for b in branches],
        "orders_by_status": {str(k or "غير محدد"): int(v or 0) for k, v in orders_by_status.items()} if scope["orders"] else _restricted(),
        "reserved_stock_lines": len(reserved) if scope["inventory"] else None,
        "recent_30d_sales": _money(recent_sales) if recent_sales is not None else None,
        "negative_margin_count": int(negative_margin_count or 0) if negative_margin_count is not None else None,
        "known_accounting_rules": [
            "الجرد الفعلي يشمل الطلبات بحالة تم الطلب وجاري الشحن، والقابل للبيع = الجرد الفعلي - المحجوز.",
            "سعر المنتج داخل الفاتورة يجب أن يبقى سعر بيع المنتج المخزني ولا يتغير بسبب أجرة التوصيل.",
            "أجرة التوصيل لا تُحسب ولا تُخصم تلقائياً؛ تُدخل يدوياً عند التسديد وتُسجّل كمصروف.",
            "مستحقات شركات النقل ذمم مدينة لصالح الشركة وليست ديناً على الشركة.",
            "حركة الصندوق اليدوية بدون ملاحظة/سبب واضحة علامة خطأ إدخال.",
            "أي تنفيذ تعديل يحتاج خطة وموافقة أدمن، ولا ينفذ GPT مباشرة.",
        ],
        "financial": financial_data,
    }


def _reserved_stock_map() -> dict[tuple[int, int], int]:
    """Reserved physical stock by (branch_id, product_id) for ordered/in-shipping orders."""
    default_branch = get_default_branch()
    default_branch_id = default_branch.id if default_branch else None
    rows = (
        db.session.query(
            func.coalesce(OrderItem.fulfillment_branch_id, Invoice.branch_id).label("branch_id"),
            OrderItem.product_id,
            func.coalesce(func.sum(OrderItem.quantity), 0).label("qty"),
        )
        .join(Invoice, Invoice.id == OrderItem.invoice_id)
        .filter(Invoice.status.in_(RESERVED_ORDER_STATUSES))
        .filter(OrderItem.product_id.isnot(None))
        .filter(exclude_delivery_fee_items(OrderItem))
        .group_by(func.coalesce(OrderItem.fulfillment_branch_id, Invoice.branch_id), OrderItem.product_id)
        .all()
    )
    out: dict[tuple[int, int], int] = {}
    for row in rows:
        branch_id = int(row.branch_id or default_branch_id or 0)
        if not branch_id:
            continue
        out[(branch_id, int(row.product_id))] = int(row.qty or 0)
    return out


_QUERY_STOP_WORDS = {
    "شنو", "ليش", "هذا", "هاي", "هسه", "اكو", "اكو", "على", "عن", "من", "الى", "إلى", "في", "بي", "بيه",
    "مال", "ممكن", "تقرير", "تحليل", "حلل", "شوف", "تاكد", "تأكد", "نقص", "زايد", "فرع", "شركة", "قطعة", "قطعه",
    "قطع", "شاشة", "الشاشة", "شاشه", "الشاشه", "شاشات", "الشاشات", "موديل", "حجم", "جرد", "مخزون", "طلبات", "طلب", "حالة", "ضروري", "راجع",
    "صار", "صارت", "النوع", "نوع", "طلبته", "منه", "بفرع", "بالفرع", "بشاشه", "بالشاشه", "موجود", "بس",
    "زياده", "زيادة", "زائد", "الزياده", "الزيادة", "ليشكو", "شكد", "شكتلة", "شنطاني",
    "شوفي", "شوفلي", "وحده", "واحده", "منتج", "المنتج", "للمنتج", "لمنتج",
}


def _normalize_query_token(raw: str) -> str:
    term = (raw or "").strip().lower()
    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ى": "ي",
        "ة": "ه",
    }
    for old, new in replacements.items():
        term = term.replace(old, new)
    for prefix in ("بال", "وال", "لل", "ال"):
        if term.startswith(prefix) and len(term) > len(prefix) + 2:
            term = term[len(prefix):]
            break
    for prefix in ("ب", "ل", "و"):
        if term.startswith(prefix) and len(term) > 4:
            term = term[1:]
            break
    return term


def _query_terms(message: str) -> list[str]:
    text_value = re.sub(r"[^\w\u0600-\u06FF]+", " ", message or "").strip()
    terms: list[str] = []
    for raw in text_value.split():
        term = _normalize_query_token(raw)
        if len(term) < 2 or term in _QUERY_STOP_WORDS:
            continue
        if term not in terms:
            terms.append(term)
    return terms[:10]


def _collect_query_evidence(message: str, scope: dict | None = None) -> dict:
    """Collect concrete DB rows related to the user's question.

    This keeps the assistant grounded: product/branch stock, reserved orders,
    recent invoices, and inventory movements are sent as evidence instead of
    letting GPT infer from a high-level summary.
    """
    scope = scope or _assistant_read_scope(None)
    evidence: dict[str, Any] = {"terms": _query_terms(message)}
    if not evidence["terms"]:
        return evidence

    if not scope.get("inventory"):
        evidence["inventory"] = _restricted("تفاصيل المخزون تحتاج صلاحية المخزون.")
        return evidence

    terms = evidence["terms"]
    branches = Branch.query.order_by(Branch.name.asc()).all()
    branch_matches = []
    branch_term_ids: set[str] = set()
    for branch in branches:
        haystack = _normalize_query_token(f"{branch.name or ''} {branch.code or ''}")
        matched = [term for term in terms if term in haystack]
        if matched:
            branch_matches.append(branch)
            branch_term_ids.update(matched)
    if not branch_matches and len(branches) == 1:
        branch_matches = branches

    target_branch_ids = [branch.id for branch in branch_matches]
    product_terms = [term for term in terms if term not in branch_term_ids]
    product_filters = []
    for term in product_terms:
        like = f"%{term}%"
        product_filters.extend(
            [
                Product.name.ilike(like),
                Product.sku.ilike(like),
                Product.barcode.ilike(like),
            ]
        )
    product_candidates = (
        Product.query.filter(or_(*product_filters))
        .order_by(Product.active.desc(), Product.name.asc())
        .limit(20)
        .all()
        if product_filters
        else []
    )

    scored_products: list[tuple[int, Product]] = []
    for product in product_candidates:
        haystack = _normalize_query_token(f"{product.name or ''} {product.sku or ''} {product.barcode or ''}")
        score = 0
        for term in product_terms:
            if term in haystack:
                score += 2 if term.isdigit() else 3
        if score:
            scored_products.append((score, product))
    scored_products.sort(key=lambda item: (-item[0], item[1].name or ""))
    if scored_products:
        best_score = scored_products[0][0]
        if best_score >= 5:
            products = [product for score, product in scored_products if score == best_score][:3]
        else:
            products = [product for _score, product in scored_products[:5]]
    else:
        products = []

    reserved = _reserved_stock_map()
    product_rows = []
    for product in products:
        branch_rows = []
        stock_query = BranchStock.query.filter_by(product_id=product.id)
        if target_branch_ids:
            stock_query = stock_query.filter(BranchStock.branch_id.in_(target_branch_ids))
        stock_rows = stock_query.order_by(BranchStock.branch_id.asc()).all()
        stock_by_branch_id = {row.branch_id: row for row in stock_rows}
        visible_branches = branch_matches or branches
        for branch in visible_branches:
            row = stock_by_branch_id.get(branch.id)
            qty = int(row.quantity or 0) if row else 0
            reserved_qty = int(reserved.get((branch.id, product.id), 0) or 0)
            branch_rows.append(
                {
                    "branch_id": branch.id,
                    "branch_name": branch.name,
                    "system_qty": qty,
                    "reserved_ordered_or_shipping": reserved_qty,
                    "salable_qty": qty - reserved_qty,
                    "opening_stock": int(row.opening_stock or 0) if row else 0,
                }
            )
        recent_orders = []
        if scope.get("orders"):
            order_branch_expr = func.coalesce(OrderItem.fulfillment_branch_id, Invoice.branch_id)
            order_query = (
                db.session.query(OrderItem, Invoice, Branch)
                .join(Invoice, Invoice.id == OrderItem.invoice_id)
                .outerjoin(Branch, Branch.id == order_branch_expr)
                .filter(OrderItem.product_id == product.id)
                .filter(Invoice.status.in_(["تم الطلب", "جاري الشحن", "تم التوصيل", "واصل", "تم التسليم"]))
            )
            if target_branch_ids:
                order_query = order_query.filter(order_branch_expr.in_(target_branch_ids))
            order_rows = order_query.order_by(Invoice.created_at.desc(), Invoice.id.desc()).limit(10).all()
            for item, invoice, branch in order_rows:
                recent_orders.append(
                    {
                        "invoice_id": invoice.id,
                        "status": invoice.status,
                        "payment_status": invoice.payment_status,
                        "branch_id": (branch.id if branch else None),
                        "branch_name": (branch.name if branch else "غير محدد"),
                        "quantity": int(item.quantity or 0),
                        "unit_price": int(item.price or 0),
                        "line_total": int(item.total or 0),
                        "created_at": invoice.created_at.isoformat() if invoice.created_at else None,
                        "shipping_barcode": invoice.shipping_barcode or "",
                    }
                )
        movements = []
        movement_branch_id = branch_matches[0].id if len(branch_matches) == 1 else None
        try:
            raw_movements = get_product_inventory_movements(product.id, branch_id=movement_branch_id)
            for movement in raw_movements[-8:]:
                movements.append(
                    {
                        "date": str(movement.get("date") or ""),
                        "type": movement.get("type_ar") or movement.get("type") or "",
                        "in": int(movement.get("quantity_in") or 0),
                        "out": int(movement.get("quantity_out") or 0),
                        "balance_after": int(movement.get("balance_after") or 0),
                        "reference": f"{movement.get('reference_type') or ''}#{movement.get('reference_id') or ''}",
                        "description": movement.get("description") or "",
                    }
                )
        except Exception as exc:
            movements.append({"error": str(exc)})

        product_rows.append(
            {
                "product_id": product.id,
                "name": product.name,
                "sku": product.sku or "",
                "barcode": product.barcode or "",
                "active": bool(product.active),
                "total_quantity": int(product.quantity or 0),
                "buy_price": int(product.buy_price or 0),
                "sale_price": int(product.sale_price or 0),
                "branch_stock": branch_rows,
                "recent_orders": recent_orders,
                "recent_movements": movements,
            }
        )

    if product_rows:
        evidence["products"] = product_rows
        evidence["branch_matches"] = [{"id": b.id, "name": b.name, "code": b.code} for b in branch_matches]
        evidence["scope_note"] = (
            "، ".join(b.name for b in branch_matches if b.name)
            if branch_matches
            else "كل الفروع"
        )
    else:
        low_rows = (
            db.session.query(Product, BranchStock, Branch)
            .join(BranchStock, BranchStock.product_id == Product.id)
            .join(Branch, Branch.id == BranchStock.branch_id)
            .filter(Product.active == True)  # noqa: E712
            .filter(BranchStock.quantity <= BranchStock.low_stock_threshold)
            .order_by(BranchStock.quantity.asc(), Product.name.asc())
            .limit(12)
            .all()
        )
        evidence["low_stock_samples"] = [
            {
                "product_id": product.id,
                "product_name": product.name,
                "branch_id": branch.id,
                "branch_name": branch.name,
                "qty": int(stock.quantity or 0),
                "threshold": int(stock.low_stock_threshold or 0),
            }
            for product, stock, branch in low_rows
        ]
    return evidence


def _normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    aliases = {
        "branch_id": {"رقم الفرع", "branch id", "branch_id"},
        "product_id": {"رقم المنتج", "product id", "product_id"},
        "branch_name": {"الفرع", "branch"},
        "product_name": {"المنتج", "product", "اسم المنتج"},
        "sku": {"sku", "SKU"},
        "barcode": {"الباركود", "barcode"},
        "system_qty": {"كمية النظام بالفرع", "كمية النظام", "system qty", "system_qty"},
        "actual_qty": {"الكمية الفعلية", "الجرد الفعلي", "actual qty", "actual_qty"},
        "note": {"ملاحظة الجرد", "ملاحظة", "note"},
    }
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40)):
        cells = {_normalize_text(cell.value): cell.column for cell in row if _normalize_text(cell.value)}
        lower_cells = {key.lower(): col for key, col in cells.items()}
        found: dict[str, int] = {}
        for field, names in aliases.items():
            for name in names:
                if name in cells:
                    found[field] = cells[name]
                    break
                if name.lower() in lower_cells:
                    found[field] = lower_cells[name.lower()]
                    break
        if "product_id" in found and "actual_qty" in found:
            return row[0].row, found
    raise ValueError("لم أجد أعمدة الجرد المطلوبة داخل ملف Excel")


def parse_inventory_audit_workbook(file_path: str) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=False, read_only=True)
    ws = wb["الجرد"] if "الجرد" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, columns = _find_header_row(ws)
    rows: list[dict] = []
    errors: list[str] = []

    def cell_value(row, field):
        idx = columns.get(field)
        if not idx:
            return None
        return row[idx - 1].value

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        product_id_raw = cell_value(row, "product_id")
        actual_raw = cell_value(row, "actual_qty")
        if product_id_raw in (None, "") and actual_raw in (None, ""):
            continue
        try:
            product_id = int(float(product_id_raw))
        except (TypeError, ValueError):
            continue
        branch_id = None
        branch_id_raw = cell_value(row, "branch_id")
        try:
            branch_id = int(float(branch_id_raw)) if branch_id_raw not in (None, "") else None
        except (TypeError, ValueError):
            branch_id = None
        try:
            actual_qty = int(float(actual_raw)) if actual_raw not in (None, "") else None
        except (TypeError, ValueError):
            errors.append(f"كمية فعلية غير صالحة للمنتج {product_id}: {actual_raw}")
            continue
        try:
            system_qty = int(float(cell_value(row, "system_qty") or 0))
        except (TypeError, ValueError):
            system_qty = 0
        rows.append(
            {
                "branch_id": branch_id,
                "product_id": product_id,
                "branch_name": _normalize_text(cell_value(row, "branch_name")),
                "product_name": _normalize_text(cell_value(row, "product_name")),
                "sku": _normalize_text(cell_value(row, "sku")),
                "barcode": _normalize_text(cell_value(row, "barcode")),
                "system_qty_file": system_qty,
                "actual_qty": actual_qty,
                "note": _normalize_text(cell_value(row, "note")),
            }
        )
    return {"sheet": ws.title, "rows": rows, "errors": errors, "row_count": len(rows)}


def save_uploaded_file(file_storage: FileStorage, *, employee_id: int | None, session_id: int | None = None) -> AIUploadedFile:
    if not file_storage or not file_storage.filename:
        raise ValueError("لم يتم اختيار ملف")
    ext = Path(file_storage.filename).suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise ValueError("ارفع ملف Excel بصيغة .xlsx أو .xlsm")

    upload_root = Path(current_app.root_path) / "uploads" / "ai_assistant" / datetime.utcnow().strftime("%Y%m%d")
    upload_root.mkdir(parents=True, exist_ok=True)
    original = file_storage.filename
    safe_name = secure_filename(original) or f"audit{ext}"
    stored_name = f"{datetime.utcnow().strftime('%H%M%S')}_{uuid.uuid4().hex[:8]}_{safe_name}"
    stored_path = upload_root / stored_name
    file_storage.save(stored_path)

    uploaded = AIUploadedFile(
        session_id=session_id,
        employee_id=employee_id,
        original_name=original,
        stored_path=str(stored_path),
        size_bytes=stored_path.stat().st_size,
        file_type="inventory_audit",
    )
    try:
        parsed = parse_inventory_audit_workbook(str(stored_path))
        preview_rows = parsed["rows"][:8]
        uploaded.status = "parsed"
        uploaded.set_preview(
            {
                "row_count": parsed["row_count"],
                "errors": parsed["errors"][:10],
                "rows": preview_rows,
            }
        )
    except Exception as exc:
        uploaded.status = "error"
        uploaded.error_message = str(exc)
        uploaded.set_preview({"error": str(exc)})
    db.session.add(uploaded)
    db.session.flush()
    _log_tool(
        "excel.parse_inventory_audit",
        session_id=session_id,
        employee_id=employee_id,
        input_data={"filename": original},
        output_data=uploaded.get_preview(),
        status="success" if uploaded.status == "parsed" else "error",
        error=uploaded.error_message,
    )
    return uploaded


def _product_label(product: Product | None, fallback: str = "") -> str:
    return (product.name if product else fallback) or "منتج غير معروف"


def _branch_label(branch: Branch | None, fallback: str = "") -> str:
    return (branch.name if branch else fallback) or "فرع غير معروف"


def build_inventory_reconcile_plan(
    upload_ids: list[int],
    *,
    employee_id: int | None,
    session_id: int | None = None,
) -> tuple[AIActionPlan | None, dict]:
    ensure_product_schema()
    ensure_branch_schema()
    uploads = AIUploadedFile.query.filter(AIUploadedFile.id.in_(upload_ids)).all() if upload_ids else []
    parsed_rows: list[dict] = []
    parse_errors: list[str] = []
    for uploaded in uploads:
        if uploaded.status != "parsed":
            parse_errors.append(f"{uploaded.original_name}: {uploaded.error_message or 'غير مقروء'}")
            continue
        parsed = parse_inventory_audit_workbook(uploaded.stored_path)
        parse_errors.extend(parsed.get("errors") or [])
        parsed_rows.extend(parsed.get("rows") or [])

    if not parsed_rows:
        return None, {"errors": parse_errors or ["لا توجد أسطر جرد فعلية مقروءة"], "rows": 0}

    reserved = _reserved_stock_map()
    branches_by_id = {b.id: b for b in Branch.query.all()}
    products_by_id = {p.id: p for p in Product.query.all()}
    default_branch = get_default_branch()

    diffs: dict[tuple[int, int], dict] = {}
    issues: list[dict] = []
    for row in parsed_rows:
        product_id = int(row["product_id"])
        product = products_by_id.get(product_id)
        if not product:
            issues.append({"severity": "warning", "message": f"المنتج رقم {product_id} موجود بالملف وغير موجود بالنظام"})
            continue
        branch_id = row.get("branch_id") or (default_branch.id if default_branch else None)
        if not branch_id:
            issues.append({"severity": "critical", "message": f"لا يمكن تحديد فرع المنتج {product.name}"})
            continue
        actual_qty = row.get("actual_qty")
        if actual_qty is None:
            continue
        actual_qty = int(actual_qty)
        reserved_qty = int(reserved.get((int(branch_id), product_id), 0))
        target_salable = actual_qty - reserved_qty
        current_salable = get_branch_stock(int(branch_id), product_id)
        if target_salable < 0:
            issues.append(
                {
                    "severity": "critical",
                    "branch_id": branch_id,
                    "product_id": product_id,
                    "message": f"الحجز ({reserved_qty}) أكبر من الجرد الفعلي ({actual_qty}) للمنتج {product.name}",
                }
            )
            continue
        diff = target_salable - current_salable
        if diff:
            diffs[(int(branch_id), product_id)] = {
                "branch_id": int(branch_id),
                "product_id": product_id,
                "actual_qty": actual_qty,
                "reserved_qty": reserved_qty,
                "target_salable": target_salable,
                "current_salable": current_salable,
                "diff": diff,
                "file_note": row.get("note") or "",
            }

    plan = AIActionPlan(
        session_id=session_id,
        created_by_id=employee_id,
        title="خطة تسوية جرد المخزون",
        plan_type="inventory_reconcile",
        status="draft",
        summary="خطة مقترحة فقط. لا يتم تعديل المخزون إلا بعد موافقة الأدمن وتنفيذ الخطة.",
        risk_level="high" if issues else "medium",
    )
    plan.set_impact(
        {
            "uploads": [u.original_name for u in uploads],
            "parsed_rows": len(parsed_rows),
            "issues": issues[:80],
            "rule": "الجرد الفعلي يشمل تم الطلب وجاري الشحن؛ القابل للبيع = الجرد الفعلي - المحجوز.",
        }
    )
    db.session.add(plan)
    db.session.flush()

    positives = [dict(v) for v in diffs.values() if v["diff"] > 0]
    negatives = [dict(v) for v in diffs.values() if v["diff"] < 0]
    positives.sort(key=lambda x: (x["product_id"], -x["diff"]))
    negatives.sort(key=lambda x: (x["product_id"], x["diff"]))

    transfer_count = 0
    for pos in positives:
        needed = int(pos["diff"])
        for neg in [n for n in negatives if n["product_id"] == pos["product_id"] and int(n["diff"]) < 0]:
            if needed <= 0:
                break
            available_extra = abs(int(neg["diff"]))
            qty = min(needed, available_extra)
            if qty <= 0:
                continue
            product = products_by_id.get(pos["product_id"])
            from_branch = branches_by_id.get(neg["branch_id"])
            to_branch = branches_by_id.get(pos["branch_id"])
            item = AIActionItem(
                plan_id=plan.id,
                item_type="inventory_transfer",
                target_type="product",
                target_id=pos["product_id"],
                title=f"تحويل {qty} من {_branch_label(from_branch)} إلى {_branch_label(to_branch)}",
                description=f"نقل فرق جرد للمنتج {_product_label(product)} بدل تسوية منفصلة.",
            )
            item.set_before(
                {
                    "from_branch_stock": get_branch_stock(neg["branch_id"], pos["product_id"]),
                    "to_branch_stock": get_branch_stock(pos["branch_id"], pos["product_id"]),
                }
            )
            item.set_after(
                {
                    "from_branch_stock": get_branch_stock(neg["branch_id"], pos["product_id"]) - qty,
                    "to_branch_stock": get_branch_stock(pos["branch_id"], pos["product_id"]) + qty,
                }
            )
            item.set_payload(
                {
                    "from_branch_id": neg["branch_id"],
                    "to_branch_id": pos["branch_id"],
                    "product_id": pos["product_id"],
                    "quantity": qty,
                    "reason": "تسوية فرق جرد بواسطة مساعد Finora",
                }
            )
            db.session.add(item)
            transfer_count += 1
            needed -= qty
            pos["diff"] -= qty
            neg["diff"] += qty

    adjustment_count = 0
    for data in positives + negatives:
        adjustment = int(data["diff"])
        if adjustment == 0:
            continue
        product = products_by_id.get(data["product_id"])
        branch = branches_by_id.get(data["branch_id"])
        action = "زيادة" if adjustment > 0 else "نقص"
        item = AIActionItem(
            plan_id=plan.id,
            item_type="stock_adjustment",
            target_type="branch_stock",
            target_id=data["product_id"],
            title=f"تسوية {action} {abs(adjustment)} من {_product_label(product)} في {_branch_label(branch)}",
            description="تسوية متبقي فرق الجرد بعد احتساب المحجوز وجاري الشحن والتحويلات الممكنة.",
        )
        before_qty = get_branch_stock(data["branch_id"], data["product_id"])
        item.set_before(
            {
                "branch_stock": before_qty,
                "actual_physical": data["actual_qty"],
                "reserved_orders": data["reserved_qty"],
                "target_salable": data["target_salable"],
            }
        )
        item.set_after({"branch_stock": max(0, before_qty + adjustment)})
        item.set_payload(
            {
                "branch_id": data["branch_id"],
                "product_id": data["product_id"],
                "adjustment": adjustment,
                "reason": "تسوية فرق جرد بواسطة مساعد Finora",
                "file_note": data.get("file_note") or "",
            }
        )
        db.session.add(item)
        adjustment_count += 1

    if transfer_count == 0 and adjustment_count == 0:
        plan.summary = "لم أجد فروقات قابلة للتنفيذ بعد احتساب المحجوز وجاري الشحن."
    else:
        plan.summary = f"الخطة تحتوي {transfer_count} تحويل و{adjustment_count} تسوية مخزون."

    _log_tool(
        "inventory.build_reconcile_plan",
        session_id=session_id,
        plan_id=plan.id,
        employee_id=employee_id,
        input_data={"upload_ids": upload_ids},
        output_data={"items": transfer_count + adjustment_count, "issues": len(issues), "errors": parse_errors[:20]},
        mode="plan",
    )
    return plan, {"errors": parse_errors, "issues": issues, "items": transfer_count + adjustment_count}


def _extract_order_ids(message: str) -> list[int]:
    text = message or ""
    text = re.sub(r"(?:باركود|barcode)\s*[:：]?\s*[\w\-]+", " ", text, flags=re.IGNORECASE)

    ids: list[int] = []
    explicit_patterns = (
        r"#\s*(\d{1,8})",
        r"(?:طلب|الطلب|فاتورة|الفاتورة|رقم)\s*#?\s*(\d{1,8})",
    )

    for pattern in explicit_patterns:
        for raw in re.findall(pattern, text):
            try:
                value = int(raw)
            except ValueError:
                continue
            if value > 0 and value not in ids:
                ids.append(value)

    if ids:
        return ids[:25]

    for raw in re.findall(r"(?<![\w\-])(\d{1,8})(?![\w\-])", text):
        try:
            value = int(raw)
        except ValueError:
            continue
        if value > 0 and value not in ids:
            ids.append(value)
    return ids[:25]


def _detect_order_action(message: str) -> str | None:
    text = message or ""
    if any(word in text for word in ("سدد", "تسديد", "مسدد", "تحصيل", "ادفع", "دفع الطلب")):
        return "mark_paid"
    if any(word in text for word in ("الغاء", "إلغاء", "ألغي", "لغي", "ملغي", "احذف الطلب")):
        return "cancel"
    if any(word in text for word in ("ارجاع", "إرجاع", "رجع", "راجع", "مرتجع", "رجيع")):
        return "return"
    return None


def build_order_action_plan(
    message: str,
    *,
    employee_id: int | None,
    session_id: int | None = None,
) -> AIActionPlan | None:
    action = _detect_order_action(message)
    order_ids = _extract_order_ids(message)
    if not action or not order_ids:
        return None

    _ensure_audit_schema()
    orders = Invoice.query.filter(Invoice.id.in_(order_ids)).order_by(Invoice.id.asc()).all()
    if not orders:
        return None

    action_label = {
        "mark_paid": "تسديد",
        "cancel": "إلغاء",
        "return": "إرجاع",
    }.get(action, action)
    plan = AIActionPlan(
        session_id=session_id,
        created_by_id=employee_id,
        title=f"خطة {action_label} طلبات",
        plan_type="order_action",
        status="draft",
        summary=f"خطة مقترحة لـ {action_label} {len(orders)} طلب. التنفيذ يحتاج موافقة أدمن.",
        risk_level="high" if action in {"cancel", "return"} else "medium",
    )
    missing = sorted(set(order_ids) - {int(o.id) for o in orders})
    plan.set_impact(
        {
            "requested_order_ids": order_ids,
            "missing_order_ids": missing,
            "action": action,
            "note": "التنفيذ يستخدم نفس منطق الطلبات الرسمي مع سجل التحصيل/المخزون.",
        }
    )
    db.session.add(plan)
    db.session.flush()

    barcode_match = re.search(r"(?:باركود|barcode)\s*[:：]?\s*([A-Za-z0-9\\-_/]+)", message or "", flags=re.IGNORECASE)
    scanned_barcode = barcode_match.group(1).strip() if barcode_match else ""

    for order in orders:
        before = {
            "status": order.status,
            "payment_status": order.payment_status,
            "paid_amount": int(order.paid_amount or 0),
            "total": int(order.total or 0),
        }
        after = dict(before)
        if action == "mark_paid":
            after.update({"status": "تم التوصيل", "payment_status": "مسدد", "paid_amount": int(order.total or 0)})
        elif action == "cancel":
            after.update({"status": "ملغي", "payment_status": "ملغي", "paid_amount": 0})
        elif action == "return":
            after.update({"status": "مرتجع", "payment_status": "مرتجع", "paid_amount": 0})
        item = AIActionItem(
            plan_id=plan.id,
            item_type="order_update",
            target_type="invoice",
            target_id=order.id,
            title=f"{action_label} الطلب #{order.id}",
            description=f"الزبون: {order.customer_name or '-'} | المبلغ: {int(order.total or 0):,} د.ع",
        )
        item.set_before(before)
        item.set_after(after)
        item.set_payload(
            {
                "invoice_id": order.id,
                "action": action,
                "barcode": scanned_barcode,
            }
        )
        db.session.add(item)

    _log_tool(
        "orders.build_action_plan",
        session_id=session_id,
        plan_id=plan.id,
        employee_id=employee_id,
        input_data={"message": message, "order_ids": order_ids},
        output_data={"orders": len(orders), "missing": missing, "action": action},
        mode="plan",
    )
    return plan


def build_supplier_ledger_plan(*, employee_id: int | None, session_id: int | None = None) -> AIActionPlan | None:
    report = audit_and_repair_supplier_ledgers(dry_run=True).to_dict()
    differences = report.get("differences") or []
    if not differences:
        return None

    plan = AIActionPlan(
        session_id=session_id,
        created_by_id=employee_id,
        title="خطة إصلاح أرصدة الموردين",
        plan_type="supplier_ledger_fix",
        status="draft",
        summary=f"وجدت {len(differences)} مورد يحتاج تصحيح. التنفيذ يحتاج موافقة أدمن.",
        risk_level="high",
    )
    plan.set_impact(report)
    db.session.add(plan)
    db.session.flush()

    for diff in differences[:200]:
        supplier_id = int(diff["supplier_id"])
        item = AIActionItem(
            plan_id=plan.id,
            item_type="supplier_ledger_fix",
            target_type="supplier",
            target_id=supplier_id,
            title=f"تصحيح رصيد المورد {diff.get('name') or supplier_id}",
            description=(
                f"الدين {int(diff.get('current_debt') or 0):,} -> {int(diff.get('expected_debt') or 0):,}، "
                f"المدفوع {int(diff.get('current_paid') or 0):,} -> {int(diff.get('expected_paid') or 0):,}"
            ),
        )
        item.set_before(
            {
                "total_debt": int(diff.get("current_debt") or 0),
                "total_paid": int(diff.get("current_paid") or 0),
                "remaining": int(diff.get("current_remaining") or 0),
            }
        )
        item.set_after(
            {
                "total_debt": int(diff.get("expected_debt") or 0),
                "total_paid": int(diff.get("expected_paid") or 0),
                "remaining": int(diff.get("expected_remaining") or 0),
            }
        )
        item.set_payload(
            {
                "supplier_id": supplier_id,
                "expected_debt": int(diff.get("expected_debt") or 0),
                "expected_paid": int(diff.get("expected_paid") or 0),
            }
        )
        db.session.add(item)

    _log_tool(
        "suppliers.build_ledger_fix_plan",
        session_id=session_id,
        plan_id=plan.id,
        employee_id=employee_id,
        output_data={"differences": len(differences)},
        mode="plan",
    )
    return plan


def build_shipping_opening_balance_plan(
    message: str,
    *,
    employee_id: int | None,
    session_id: int | None = None,
) -> AIActionPlan | None:
    text = message or ""
    if not any(word in text for word in ("افتتاحي", "الرصيد الافتتاحي", "رصيد افتتاحي")):
        return None
    if not any(word in text for word in ("شركة النقل", "شركات النقل", "شحن", "النقل")):
        return None
    amount_match = re.search(r"(\d[\d,]{4,})", text)
    if not amount_match:
        return None
    target_amount = int(amount_match.group(1).replace(",", ""))

    company = None
    id_match = re.search(r"(?:شركة النقل|شحن|النقل)\s*(?:رقم)?\s*#?\s*(\d+)", text)
    if id_match:
        company = ShippingCompany.query.get(int(id_match.group(1)))
    if not company:
        companies = ShippingCompany.query.order_by(ShippingCompany.id.asc()).all()
        if len(companies) == 1:
            company = companies[0]
    if not company:
        return None

    plan = AIActionPlan(
        session_id=session_id,
        created_by_id=employee_id,
        title="خطة تعديل رصيد افتتاحي لشركة نقل",
        plan_type="shipping_opening_balance_fix",
        status="draft",
        summary=f"تعديل رصيد {company.name} الافتتاحي إلى {target_amount:,} د.ع كذمة مدينة لصالح الشركة.",
        risk_level="high",
    )
    plan.set_impact(
        {
            "shipping_company_id": company.id,
            "shipping_company": company.name,
            "meaning": "هذا الرصيد ذمة إلنا عند شركة النقل وليس ديناً علينا.",
        }
    )
    db.session.add(plan)
    db.session.flush()

    item = AIActionItem(
        plan_id=plan.id,
        item_type="shipping_opening_balance_fix",
        target_type="shipping_company",
        target_id=company.id,
        title=f"تعديل رصيد {company.name}",
        description=f"الرصيد الافتتاحي {int(company.opening_balance or 0):,} -> {target_amount:,} د.ع",
    )
    item.set_before({"opening_balance": int(company.opening_balance or 0)})
    item.set_after({"opening_balance": target_amount})
    item.set_payload({"shipping_company_id": company.id, "opening_balance": target_amount})
    db.session.add(item)
    _log_tool(
        "shipping.build_opening_balance_plan",
        session_id=session_id,
        plan_id=plan.id,
        employee_id=employee_id,
        input_data={"message": message},
        output_data={"shipping_company_id": company.id, "opening_balance": target_amount},
        mode="plan",
    )
    return plan


def _extract_iqd_amount(message: str) -> int:
    text_value = (message or "").replace(",", "")
    scaled = re.findall(r"(\d+(?:\.\d+)?)\s*(مليون|الف|ألف)", text_value)
    if scaled:
        value, unit = scaled[-1]
        factor = 1_000_000 if unit == "مليون" else 1_000
        return int(float(value) * factor)
    values = []
    for raw in re.findall(r"(?<![\w-])(\d{4,12})(?![\w-])", text_value):
        try:
            values.append(int(raw))
        except ValueError:
            continue
    return max(values) if values else 0


def build_shipping_duplicate_payment_plan(
    message: str,
    *,
    employee_id: int | None,
    session_id: int | None = None,
) -> tuple[AIActionPlan | None, dict]:
    text_value = message or ""
    shipping_words = ("شركة النقل", "شركة الشحن", "النقل", "الشحن")
    duplicate_words = ("مرتين", "مكرر", "مكررة", "بالغلط", "بالخطأ", "احذف وحدة", "حذف وحدة")
    if not any(word in text_value for word in shipping_words) or not any(word in text_value for word in duplicate_words):
        return None, {}

    amount = _extract_iqd_amount(text_value)
    query = ShippingPayment.query
    if amount > 0:
        query = query.filter(ShippingPayment.amount == amount)
    rows = query.order_by(ShippingPayment.created_at.desc(), ShippingPayment.id.desc()).limit(100).all()
    groups: dict[tuple, list[ShippingPayment]] = defaultdict(list)
    for payment in rows:
        key = (
            int(payment.shipping_company_id),
            int(payment.invoice_id) if payment.invoice_id else None,
            int(payment.amount or 0),
            payment.action or "",
            int(payment.treasury_account_id) if payment.treasury_account_id else None,
            payment.created_at.date().isoformat() if payment.created_at else "",
        )
        groups[key].append(payment)
    duplicate_groups = [(key, items) for key, items in groups.items() if len(items) > 1]
    if len(duplicate_groups) != 1:
        return None, {
            "matched_groups": len(duplicate_groups),
            "amount": amount,
            "message": "لم أجد تكراراً واحداً مؤكداً" if not duplicate_groups else "وجدت أكثر من مجموعة مكررة؛ حدد الشركة أو رقم الحركة",
        }

    key, payments = duplicate_groups[0]
    company = ShippingCompany.query.get(key[0])
    payment = max(payments, key=lambda row: row.id)
    restores_opening = payment.invoice_id is None and (payment.action or "").strip() == "قبض"
    plan = AIActionPlan(
        session_id=session_id,
        created_by_id=employee_id,
        title="حذف قبض مكرر من شركة نقل",
        plan_type="shipping_duplicate_payment",
        status="draft",
        summary=(
            f"حذف الحركة الأحدث #{payment.id} بمبلغ {int(payment.amount or 0):,} د.ع من "
            f"{company.name if company else 'شركة النقل'}، بعد موافقة المدير."
        ),
        risk_level="high",
    )
    plan.set_impact(
        {
            "shipping_company_id": payment.shipping_company_id,
            "shipping_company": company.name if company else "",
            "amount": int(payment.amount or 0),
            "duplicate_payment_ids": [row.id for row in payments],
            "cash_effect": -int(payment.amount or 0),
            "opening_balance_restore": int(payment.amount or 0) if restores_opening else 0,
        }
    )
    db.session.add(plan)
    db.session.flush()
    item = AIActionItem(
        plan_id=plan.id,
        item_type="shipping_payment_delete_duplicate",
        target_type="shipping_payment",
        target_id=payment.id,
        title=f"حذف حركة القبض المكررة #{payment.id}",
        description=f"الإبقاء على الحركة #{min(row.id for row in payments)} وحذف الأحدث فقط.",
    )
    item.set_before(
        {
            "shipping_company_id": payment.shipping_company_id,
            "invoice_id": payment.invoice_id,
            "amount": int(payment.amount or 0),
            "action": payment.action or "",
            "treasury_account_id": payment.treasury_account_id,
            "duplicate_count": len(payments),
            "opening_balance": int(company.opening_balance or 0) if company else None,
        }
    )
    item.set_after({"deleted": True, "remaining_duplicate_count": len(payments) - 1})
    item.set_payload(
        {
            "payment_id": payment.id,
            "duplicate_payment_ids": [row.id for row in payments],
            "restore_opening_balance": restores_opening,
        }
    )
    db.session.add(item)
    _log_tool(
        "shipping.build_duplicate_payment_plan",
        session_id=session_id,
        plan_id=plan.id,
        employee_id=employee_id,
        input_data={"amount": amount},
        output_data={"payment_id": payment.id, "duplicates": len(payments)},
        mode="plan",
    )
    return plan, {"matched_groups": 1, "payment_id": payment.id, "duplicate_ids": [row.id for row in payments]}


def _asset_category_for_name(name: str) -> FixedAssetCategory | None:
    seed_default_categories()
    category_hint = "سيارات" if any(word in name for word in ("سيارة", "سياره")) else "أجهزة ومعدات"
    return (
        FixedAssetCategory.query.filter(FixedAssetCategory.name.ilike(f"%{category_hint}%")).first()
        or FixedAssetCategory.query.filter_by(is_active=True).order_by(FixedAssetCategory.id.asc()).first()
    )


def build_fixed_asset_action_plan(
    message: str,
    *,
    employee_id: int | None,
    session_id: int | None = None,
) -> tuple[AIActionPlan | None, dict]:
    text_value = message or ""
    if not any(word in text_value for word in ("أصل", "الاصل", "الأصل", "سيارة", "سياره", "راوتر", "لابتوب", "حاسبة")):
        return None, {}

    pending_items: list[dict] = []
    non_cash = any(
        phrase in text_value
        for phrase in ("ما ياخذ من الصندوق", "ماياخذ من الصندوق", "ماريده ياخذ من الصندوق", "بدون سحب من الصندوق")
    )
    if non_cash and any(word in text_value for word in ("سيارة", "سياره")):
        candidates = FixedAsset.query.filter(
            or_(FixedAsset.name.ilike("%سيارة%"), FixedAsset.name.ilike("%سياره%")),
            FixedAsset.acquisition_journal_entry_id.is_(None),
        ).order_by(FixedAsset.id.desc()).all()
        if len(candidates) == 1:
            asset = candidates[0]
            pending_items.append({"kind": "funding", "asset": asset})
        elif len(candidates) != 1:
            return None, {"message": "حدد أصل السيارة بالاسم أو الكود" if candidates else "لم أجد أصل سيارة غير مرحّل"}

    purchased = any(word in text_value for word in ("اشتريت", "شراء", "شريت"))
    amount = _extract_iqd_amount(text_value)
    asset_name = ""
    for keyword, label in (("راوتر", "راوتر"), ("لابتوب", "لابتوب"), ("حاسبة", "حاسبة")):
        if keyword in text_value:
            asset_name = label
            break
    if purchased and asset_name and amount > 0:
        category = _asset_category_for_name(asset_name)
        if not category:
            return None, {"message": "لا يوجد تصنيف أصول نشط"}
        pending_items.append(
            {
                "kind": "create",
                "name": asset_name,
                "amount": amount,
                "category": category,
                "payment_method": "cash" if any(word in text_value for word in ("الصندوق", "نقد", "كاش", "سحب")) else "credit",
            }
        )

    if not pending_items:
        return None, {"message": "أحتاج اسم الأصل، المبلغ، وهل الدفع من الصندوق أو آجل"}

    plan = AIActionPlan(
        session_id=session_id,
        created_by_id=employee_id,
        title="ترتيب الأصول الثابتة",
        plan_type="fixed_asset_actions",
        status="draft",
        summary=f"خطة من {len(pending_items)} إجراء على الأصول، ولا تنفذ إلا بعد فحص وموافقة المدير.",
        risk_level="high",
    )
    db.session.add(plan)
    db.session.flush()
    cash_delta = 0
    for pending in pending_items:
        if pending["kind"] == "funding":
            asset = pending["asset"]
            item = AIActionItem(
                plan_id=plan.id,
                item_type="fixed_asset_set_capital_and_post",
                target_type="fixed_asset",
                target_id=asset.id,
                title=f"ترحيل {asset.name} بدون سحب من الصندوق",
                description=f"ترحيل الأصل كإضافة مالك/رصيد افتتاحي بقيمة {int(asset.total_cost or 0):,} د.ع، بدون صندوق وبدون دين مورد.",
            )
            item.set_before(
                {
                    "status": asset.status,
                    "payment_method": asset.payment_method,
                    "paid_amount": int(asset.paid_amount or 0),
                    "credit_amount": int(asset.credit_amount or 0),
                    "total_cost": int(asset.total_cost or 0),
                    "posted": bool(asset.acquisition_journal_entry_id),
                }
            )
            item.set_after(
                {
                    "status": "active",
                    "payment_method": "capital",
                    "paid_amount": 0,
                    "credit_amount": 0,
                    "cash_effect": 0,
                }
            )
            item.set_payload({"asset_id": asset.id})
        else:
            amount = int(pending["amount"])
            payment_method = pending["payment_method"]
            cash_delta -= amount if payment_method == "cash" else 0
            item = AIActionItem(
                plan_id=plan.id,
                item_type="fixed_asset_create_and_post",
                target_type="fixed_asset",
                title=f"إضافة وترحيل أصل {pending['name']}",
                description=(
                    f"شراء {pending['name']} بقيمة {amount:,} د.ع "
                    + ("وسحبها من الصندوق." if payment_method == "cash" else "كشراء آجل بدون سحب من الصندوق.")
                ),
            )
            item.set_before({"matching_assets": 0})
            item.set_after({"name": pending["name"], "total_cost": amount, "payment_method": payment_method, "posted": True})
            item.set_payload(
                {
                    "name": pending["name"],
                    "category_id": pending["category"].id,
                    "purchase_price": amount,
                    "purchase_date": business_now_local().date().isoformat(),
                    "payment_method": payment_method,
                }
            )
        db.session.add(item)
    plan.set_impact({"items": len(pending_items), "cash_effect": cash_delta, "requires_admin_approval": True})
    _log_tool(
        "assets.build_action_plan",
        session_id=session_id,
        plan_id=plan.id,
        employee_id=employee_id,
        output_data={"items": len(pending_items), "cash_effect": cash_delta},
        mode="plan",
    )
    return plan, {"items": len(pending_items), "cash_effect": cash_delta}


def _get_openai_key() -> str:
    key = os.environ.get("OPENAI_API_KEY") or os.environ.get("openai_api_key") or ""
    if key.strip():
        return key.strip()
    try:
        from flask import g
        old_tenant = getattr(g, "tenant", None)
        g.tenant = None
        from models.core.global_setting import GlobalSetting

        key = (GlobalSetting.get_setting("OPENAI_API_KEY", "") or "").strip()
        g.tenant = old_tenant
    except Exception:
        key = ""
    return key


def _assistant_response_schema() -> dict:
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "answer": {"type": "string"},
            "evidence": {"type": "array", "items": {"type": "string"}, "maxItems": 8},
            "key_points": {"type": "array", "items": {"type": "string"}, "maxItems": 8},
            "risks": {"type": "array", "items": {"type": "string"}, "maxItems": 6},
            "next_steps": {"type": "array", "items": {"type": "string"}, "maxItems": 6},
            "needs_admin_approval": {"type": "boolean"},
        },
        "required": ["answer", "evidence", "key_points", "risks", "next_steps", "needs_admin_approval"],
    }


def _safe_json_loads(text_value: str) -> dict | None:
    if not text_value:
        return None
    try:
        parsed = json.loads(text_value)
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        pass
    match = re.search(r"\{.*\}", text_value, re.DOTALL)
    if not match:
        return None
    try:
        parsed = json.loads(match.group(0))
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None


def _format_structured_ai_reply(data: dict | None, fallback: str = "") -> str:
    if not data:
        return (fallback or "").strip()
    lines: list[str] = []
    answer = _normalize_text(data.get("answer"))
    if answer:
        lines.append(answer)
    sections = [
        ("الأدلة من النظام", data.get("evidence")),
        ("النقاط المهمة", data.get("key_points")),
        ("المخاطر", data.get("risks")),
        ("الخطوات المقترحة", data.get("next_steps")),
    ]
    for title, values in sections:
        if not isinstance(values, list) or not values:
            continue
        clean_values = [_normalize_text(v) for v in values if _normalize_text(v)]
        if not clean_values:
            continue
        lines.append(f"\n{title}:")
        lines.extend(f"- {value}" for value in clean_values[:8])
    if data.get("needs_admin_approval"):
        lines.append("\nأي تعديل فعلي يحتاج خطة وموافقة أدمن قبل التنفيذ.")
    return "\n".join(lines).strip() or (fallback or "").strip()


_AR_WEEKDAYS = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6: "الأحد"}

_TOOL_LOOP_MAX_ROUNDS = 6
_TOOL_OUTPUT_MAX_CHARS = 14000


def _assistant_system_prompt() -> str:
    now = business_now_local()
    weekday = _AR_WEEKDAYS.get(now.weekday(), "")
    return (
        "أنت المساعد المالي لنظام Finora المحاسبي. جاوب بالعربية العراقية بأسلوب محاسب محترف ومختصر.\n"
        f"الوقت المحلي الحالي في بغداد: {weekday} {now.strftime('%Y-%m-%d %H:%M')}.\n"
        "\n"
        "قواعد أساسية:\n"
        "1. عندك أدوات (tools) تقرأ البيانات الحقيقية من قاعدة بيانات الشركة. لأي سؤال عن أرقام "
        "(صندوق، مبيعات، أرباح، مصاريف، ديون، مخزون، زبائن، موردين، شحن) لازم تستدعي الأداة المناسبة "
        "وتجاوب من نتائجها الفعلية. لا تجاوب من الذاكرة ولا تخمّن أرقاماً أبداً.\n"
        "2. حوّل العبارات الزمنية إلى تواريخ فعلية اعتماداً على الوقت المحلي أعلاه: "
        "«اليوم» و«من الصبح لهسة» = تاريخ اليوم، «أمس» = اليوم السابق، «هذا الأسبوع» من السبت، "
        "«هذا الشهر» من أول الشهر. ومرر التواريخ للأدوات بصيغة YYYY-MM-DD. "
        "اذكر في جوابك الفترة التي حسبتها بوضوح.\n"
        "3. اكتب المبالغ بأرقام مفصولة بالفواصل متبوعة بـ «د.ع» (مثال: 1,250,000 د.ع).\n"
        "4. إذا الأداة رجعت error أو restricted أو ما رجعت بيانات، قلها صراحة ولا تعوّضها بتخمين.\n"
        "5. لا تخترع اسم منتج أو فرع أو زبون أو رقم طلب غير موجود في نتائج الأدوات.\n"
        "6. لا تدّعي تنفيذ أي تعديل. أنت للقراءة والتحليل فقط؛ أي تعديل فعلي يكون بخطة تنفيذ وموافقة أدمن.\n"
        "7. للمقارنات (اليوم مقابل أمس، أسبوع مقابل أسبوع) استدعِ الأداة مرتين بفترتين مختلفتين واعرض الفرق.\n"
        "8. مهم جداً: جاوب حصراً على آخر سؤال كتبه المستخدم. المحادثة السابقة للسياق فقط؛ "
        "لا تعيد جواب سؤال قديم ولا تكمل على موضوع سابق إلا إذا السؤال الأخير يشير له صراحة.\n"
        "9. إذا السؤال غامض أو ما عندك أداة تجاوب عليه، قل ذلك بوضوح واذكر شنو تقدر تجاوب عليه، "
        "بدل ما تجاوب على شي ثاني.\n"
        "10. افهم أوامر اللهجة العراقية مثل: رتبلي، صلّح، احذف وحدة، رجعها، ماريده ياخذ من الصندوق، "
        "وسحبها من الصندوق. إذا الطلب تعديل، اشرح خطة التنفيذ التي أنشأها النظام ولا تقل إنه تم التنفيذ.\n"
        "11. عند سؤال (ليش ما مبين؟) ابحث عن السجل ورقم الحركة وتاريخ تسجيلها، وميّز بين تاريخ إنشاء الطلب أو الكشف "
        "وتاريخ حركة التحصيل التي يعتمدها التقرير اليومي.\n"
        "\n"
        "قواعد Finora المحاسبية الخاصة:\n"
        "- الجرد الفعلي يشمل الطلبات بحالة «تم الطلب» و«جاري الشحن»؛ القابل للبيع = كمية النظام - المحجوز.\n"
        "- سعر المنتج في الفاتورة لا يتغير بسبب أجرة التوصيل.\n"
        "- أجرة التوصيل تُدخل يدوياً عند التسديد فقط ولا تُخصم تلقائياً.\n"
        "- مستحقات شركات النقل ذمم مدينة لصالحنا (إلنا عندهم) وليست ديناً علينا.\n"
        "- ديون الموردين التزامات علينا؛ ذمم الزبائن أصول إلنا.\n"
        "- حركة صندوق يدوية بلا ملاحظة/سبب تعتبر خطأ إدخال محتمل يستحق التنويه.\n"
        "\n"
        "بعد ما تجمع البيانات اللازمة، أرجع الجواب النهائي حصراً كـ JSON بهذا الشكل:\n"
        '{"answer": "الجواب المباشر بالأرقام", "evidence": ["أرقام حقيقية من نتائج الأدوات"], '
        '"key_points": ["نقاط مهمة"], "risks": ["مخاطر أو ملاحظات"], "next_steps": ["خطوات مقترحة"], '
        '"needs_admin_approval": false}'
    )


def _serialize_tool_output(data: dict) -> str:
    try:
        text = json.dumps(data, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        text = json.dumps({"error": "تعذر تحويل نتيجة الأداة"}, ensure_ascii=False)
    if len(text) > _TOOL_OUTPUT_MAX_CHARS:
        text = text[:_TOOL_OUTPUT_MAX_CHARS] + '... (تم اقتصاص النتيجة لكبر حجمها)"'
    return text


def _call_openai_narrative(
    message: str,
    snapshot: dict,
    local_findings: dict | None = None,
    *,
    scope: dict | None = None,
    history: list[dict] | None = None,
    session_id: int | None = None,
    employee_id: int | None = None,
    period_hint: dict | None = None,
) -> tuple[bool, str]:
    key = _get_openai_key()
    if not key:
        return False, ""
    try:
        import openai
    except ImportError:
        return False, ""

    scope = scope or _assistant_read_scope(employee_id)
    model = os.environ.get("FINORA_AI_MODEL") or os.environ.get("OPENAI_ANALYSIS_MODEL") or "gpt-4o-mini"
    tools = get_tool_definitions(scope)

    context_payload = {
        "system_snapshot": snapshot,
        "local_findings": {k: v for k, v in (local_findings or {}).items() if k != "snapshot"},
    }
    if period_hint:
        context_payload["ui_period_hint"] = period_hint

    messages: list[dict] = [{"role": "system", "content": _assistant_system_prompt()}]
    for item in history or []:
        messages.append(item)
    messages.append(
        {
            "role": "system",
            "content": (
                "سياق النظام الحالي (للاستئناس فقط، ولا يغني عن استدعاء الأدوات للأرقام الدقيقة):\n"
                + json.dumps(context_payload, ensure_ascii=False, default=str)
            ),
        }
    )
    messages.append({"role": "user", "content": message})

    try:
        client = openai.OpenAI(api_key=key)
    except Exception as exc:
        current_app.logger.warning("OpenAI client init failed: %s", exc)
        return False, ""

    try:
        loop_kwargs: dict[str, Any] = {"model": model, "max_tokens": 1800, "timeout": 45}
        if tools:
            loop_kwargs["tools"] = tools
            loop_kwargs["tool_choice"] = "auto"
        for _round in range(_TOOL_LOOP_MAX_ROUNDS):
            response = client.chat.completions.create(messages=messages, **loop_kwargs)
            choice = response.choices[0] if response.choices else None
            if not choice or not choice.message:
                break
            msg = choice.message
            tool_calls = getattr(msg, "tool_calls", None) or []
            if not tool_calls:
                text = msg.content or ""
                return True, _format_structured_ai_reply(_safe_json_loads(text), text)

            messages.append(
                {
                    "role": "assistant",
                    "content": msg.content or None,
                    "tool_calls": [
                        {
                            "id": tc.id,
                            "type": "function",
                            "function": {"name": tc.function.name, "arguments": tc.function.arguments or "{}"},
                        }
                        for tc in tool_calls
                    ],
                }
            )
            for tc in tool_calls:
                tool_name = tc.function.name
                try:
                    tool_args = json.loads(tc.function.arguments or "{}")
                    if not isinstance(tool_args, dict):
                        tool_args = {}
                except (TypeError, json.JSONDecodeError):
                    tool_args = {}
                result = execute_tool(tool_name, tool_args, scope)
                _log_tool(
                    f"ai.{tool_name}",
                    session_id=session_id,
                    employee_id=employee_id,
                    input_data=tool_args,
                    output_data={"preview": _serialize_tool_output(result)[:500]},
                    mode="read",
                    status="error" if isinstance(result, dict) and result.get("error") else "success",
                    error=str(result.get("error")) if isinstance(result, dict) and result.get("error") else None,
                )
                messages.append(
                    {
                        "role": "tool",
                        "tool_call_id": tc.id,
                        "content": _serialize_tool_output(result if isinstance(result, dict) else {"result": result}),
                    }
                )

        # تجاوز عدد الجولات: اطلب خلاصة نهائية بدون أدوات
        messages.append(
            {
                "role": "user",
                "content": "أعطِ الجواب النهائي الآن كـ JSON حسب المخطط، بالاعتماد على نتائج الأدوات أعلاه فقط.",
            }
        )
        response = client.chat.completions.create(
            model=model,
            messages=messages,
            max_tokens=1600,
            timeout=45,
        )
        choice = response.choices[0] if response.choices else None
        text = choice.message.content if choice and choice.message else ""
        return True, _format_structured_ai_reply(_safe_json_loads(text or ""), text or "")
    except Exception as exc:
        current_app.logger.warning("OpenAI tool-calling narrative failed: %s", exc)
        return False, ""


def _chat_history_messages(chat: AIChatSession, limit: int = 12) -> list[dict]:
    """آخر رسائل الجلسة (ذاكرة المحادثة) بصيغة OpenAI messages."""
    rows = (
        AIChatMessage.query.filter_by(session_id=chat.id)
        .filter(AIChatMessage.role.in_(["user", "assistant"]))
        .order_by(AIChatMessage.id.desc())
        .limit(limit)
        .all()
    )
    history: list[dict] = []
    for row in reversed(rows):
        content = (row.content or "").strip()
        if not content:
            continue
        # نقتصر ردود المساعد السابقة حتى لا تطغى على السؤال الأخير
        max_len = 700 if row.role == "assistant" else 1200
        history.append({"role": row.role, "content": content[:max_len]})
    return history


def handle_chat_send(
    *,
    employee_id: int | None,
    message: str,
    session_id: int | None = None,
    upload_ids: list[int] | None = None,
    period: str | None = None,
    analysis_type: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> dict:
    chat = get_or_create_chat_session(employee_id, session_id)
    history = _chat_history_messages(chat)
    add_chat_message(chat, "user", message, {"upload_ids": upload_ids or []})
    upload_ids = [int(x) for x in (upload_ids or []) if str(x).isdigit()]
    period_hint = {
        k: v
        for k, v in {
            "type": analysis_type,
            "period": period,
            "date_from": date_from,
            "date_to": date_to,
        }.items()
        if v
    } or None

    scope = _assistant_read_scope(employee_id)
    snapshot = collect_system_snapshot(employee_id=employee_id)
    local_findings: dict[str, Any] = {"snapshot": snapshot}
    query_evidence = _collect_query_evidence(message or "", scope)
    if query_evidence.get("products") or query_evidence.get("low_stock_samples"):
        local_findings["query_evidence"] = query_evidence
    plan = None

    wants_inventory_plan = bool(upload_ids) or any(word in (message or "") for word in ("جرد", "فروقات", "فرق", "حول", "نقص", "زايد"))
    if wants_inventory_plan and upload_ids and scope["inventory"]:
        plan, plan_meta = build_inventory_reconcile_plan(upload_ids, employee_id=employee_id, session_id=chat.id)
        local_findings["inventory_plan"] = plan_meta
    elif wants_inventory_plan and upload_ids:
        local_findings["inventory_plan"] = {"restricted": True, "message": "تحليل وتنفيذ الجرد يحتاج صلاحية إدارة المخزون."}

    order_plan = build_order_action_plan(message, employee_id=employee_id, session_id=chat.id) if scope["order_manage"] else None
    if order_plan:
        plan = order_plan
        local_findings["order_plan"] = {"id": order_plan.id, "items": len(order_plan.items)}
    elif _detect_order_action(message) and not scope["order_manage"]:
        local_findings["order_plan"] = {"restricted": True, "message": "اقتراح تسديد/إلغاء/إرجاع الطلب يحتاج صلاحية إدارة الطلبات."}

    if any(word in (message or "") for word in ("مورد", "موردين", "الموردين", "ديون المورد")) and scope["suppliers"]:
        supplier_plan = build_supplier_ledger_plan(employee_id=employee_id, session_id=chat.id)
        if supplier_plan:
            plan = supplier_plan
            local_findings["supplier_plan"] = {"id": supplier_plan.id, "items": len(supplier_plan.items)}
    elif any(word in (message or "") for word in ("مورد", "موردين", "الموردين", "ديون المورد")):
        local_findings["supplier_plan"] = {"restricted": True, "message": "تحليل أرصدة الموردين يحتاج صلاحية الموردين."}

    duplicate_shipping_plan = None
    duplicate_shipping_meta = {}
    if scope["shipping"]:
        duplicate_shipping_plan, duplicate_shipping_meta = build_shipping_duplicate_payment_plan(
            message, employee_id=employee_id, session_id=chat.id
        )
    if duplicate_shipping_plan:
        plan = duplicate_shipping_plan
        local_findings["shipping_duplicate_plan"] = {
            "id": duplicate_shipping_plan.id,
            "items": len(duplicate_shipping_plan.items),
            **duplicate_shipping_meta,
        }
    elif duplicate_shipping_meta:
        local_findings["shipping_duplicate_plan"] = duplicate_shipping_meta

    shipping_plan = None
    if plan is None and scope["shipping"]:
        shipping_plan = build_shipping_opening_balance_plan(message, employee_id=employee_id, session_id=chat.id)
    if shipping_plan:
        plan = shipping_plan
        local_findings["shipping_plan"] = {"id": shipping_plan.id, "items": len(shipping_plan.items)}
    elif any(word in (message or "") for word in ("نقل", "شحن", "شركة النقل", "شركات النقل")) and not scope["shipping"]:
        local_findings["shipping_plan"] = {"restricted": True, "message": "تحليل شركات النقل يحتاج صلاحية الشحن."}

    asset_plan = None
    asset_meta = {}
    if plan is None and scope["assets"]:
        asset_plan, asset_meta = build_fixed_asset_action_plan(
            message, employee_id=employee_id, session_id=chat.id
        )
    if asset_plan:
        plan = asset_plan
        local_findings["fixed_asset_plan"] = {"id": asset_plan.id, **asset_meta}
    elif asset_meta:
        local_findings["fixed_asset_plan"] = asset_meta
    elif any(word in (message or "") for word in ("أصل", "الاصل", "الأصل", "سيارة", "سياره", "راوتر")) and not scope["assets"]:
        local_findings["fixed_asset_plan"] = {"restricted": True, "message": "إدارة الأصول تحتاج صلاحية الأصول الثابتة."}

    if any(word in (message or "") for word in ("محاسبي", "حساب", "هامش", "سالب", "تقرير", "اخطاء", "أخطاء", "صندوق", "كاش", "نقد", "فروقات")) and (scope["financial"] or scope["reports"]):
        audit = audit_accounting_integrity(limit=120)
        local_findings["accounting_audit"] = audit.get("summary", {})
        local_findings["accounting_audit_samples"] = {
            "stock_imbalances": (audit.get("stock_imbalances") or [])[:5],
            "invoice_total_mismatches": (audit.get("invoice_total_mismatches") or [])[:5],
            "negative_margin_items": (audit.get("negative_margin_items") or [])[:5],
            "status_inconsistencies": (audit.get("status_inconsistencies") or [])[:5],
        }
        _log_tool(
            "accounting.audit_integrity",
            session_id=chat.id,
            employee_id=employee_id,
            input_data={"limit": 120},
            output_data=audit.get("summary", {}),
        )
    elif any(word in (message or "") for word in ("محاسبي", "حساب", "هامش", "سالب", "تقرير", "اخطاء", "أخطاء", "صندوق", "كاش", "نقد", "فروقات")):
        local_findings["accounting_audit"] = {"restricted": True, "message": "التدقيق المحاسبي يحتاج صلاحية التقارير أو المالية."}

    ok, ai_text = _call_openai_narrative(
        message,
        snapshot,
        local_findings,
        scope=scope,
        history=history,
        session_id=chat.id,
        employee_id=employee_id,
        period_hint=period_hint,
    )
    if not ok or not (ai_text or "").strip():
        ai_text = _local_fallback_answer(message, scope, local_findings)
        evidence_text = _local_evidence_text(local_findings.get("query_evidence") or {})
        if evidence_text and "الأدلة المحلية المباشرة" not in ai_text:
            ai_text = f"{ai_text}\n{evidence_text}"
    if plan:
        ai_text = f"{ai_text}\n\nتم إنشاء خطة تنفيذ #{plan.id}: {plan.summary}"

    add_chat_message(chat, "assistant", ai_text, {"action_plan_id": plan.id if plan else None, "local_findings": local_findings})
    db.session.commit()
    return {
        "success": True,
        "session_id": chat.id,
        "reply": ai_text,
        "action_plan": _serialize_plan(plan),
        "local_findings": local_findings,
    }


def _local_fallback_answer(message: str, scope: dict, local_findings: dict) -> str:
    """رد محلي بأرقام حقيقية عند تعذر الاتصال بـ GPT (مفتاح مفقود أو خطأ اتصال)."""
    text = message or ""
    lines: list[str] = []

    def _fmt(value) -> str:
        return f"{int(value or 0):,} د.ع"

    if any(w in text for w in ("صندوق", "كاش", "نقد", "خزين", "سيولة")) and scope.get("financial"):
        data = execute_tool("get_cash_movements", {}, scope)
        if isinstance(data, dict) and not data.get("error") and not data.get("restricted"):
            period = data.get("period") or {}
            lines.append(f"حساب {data.get('account', {}).get('name', 'الصندوق')} ليوم {period.get('from', '')}:")
            lines.append(f"- رصيد بداية اليوم: {_fmt(data.get('opening_balance'))}")
            lines.append(f"- الداخل: {_fmt(data.get('total_in'))} | الخارج: {_fmt(data.get('total_out'))}")
            lines.append(f"- رصيد نهاية الفترة: {_fmt(data.get('closing_balance'))} | الرصيد الحالي: {_fmt(data.get('current_balance_now'))}")
            movements = data.get("movements") or []
            if movements:
                lines.append(f"- آخر الحركات ({min(len(movements), 5)} من {data.get('movements_count_in_period', 0)}):")
                for m in movements[-5:]:
                    lines.append(f"  • {m.get('type')} {_fmt(m.get('amount'))} — {m.get('description') or m.get('reason')}")

    if any(w in text for w in ("مبيعات", "بيعت", "مبيع")) and (scope.get("reports") or scope.get("financial")):
        data = execute_tool("get_sales_summary", {}, scope)
        if isinstance(data, dict) and not data.get("error") and not data.get("restricted"):
            lines.append(f"مبيعات اليوم: {data.get('orders_count', 0)} طلب بإجمالي {_fmt(data.get('total_sales'))} (المحصّل {_fmt(data.get('collected_cash'))}).")

    if "ربح" in text and (scope.get("financial") or scope.get("reports")):
        data = execute_tool("get_profit_summary", {}, scope)
        if isinstance(data, dict) and not data.get("error") and not data.get("restricted"):
            lines.append(
                f"ربح اليوم: مبيعات {_fmt(data.get('sales_total'))} - تكلفة {_fmt(data.get('cogs'))} - مصاريف {_fmt(data.get('expenses'))} = صافي {_fmt(data.get('net_profit'))}."
            )

    if any(w in text for w in ("مصروف", "مصاريف")) and scope.get("financial"):
        data = execute_tool("get_expenses", {}, scope)
        if isinstance(data, dict) and not data.get("error") and not data.get("restricted"):
            period = data.get("period") or {}
            lines.append(f"مصاريف الفترة {period.get('from', '')} → {period.get('to', '')}: {_fmt(data.get('total'))} ({data.get('count', 0)} مصروف).")

    if any(w in text for w in ("مورد", "موردين")) and (scope.get("suppliers") or scope.get("financial")):
        data = execute_tool("get_supplier_debts", {}, scope)
        if isinstance(data, dict) and not data.get("error") and not data.get("restricted"):
            lines.append(f"إجمالي المتبقي للموردين: {_fmt(data.get('total_remaining_debt'))} على {data.get('suppliers_count', 0)} مورد.")

    if lines:
        lines.append("\n(تعذر الاتصال بمحرك GPT، فهذه إجابة محلية مباشرة من بيانات النظام.)")
        return "\n".join(lines)
    return (
        "تعذر الاتصال بمحرك GPT حالياً. هذا ملخص محلي من النظام:\n" + _local_summary_text(local_findings)
    )


def _local_summary_text(findings: dict) -> str:
    snap = findings.get("snapshot") or {}
    products = snap.get("products") or {}
    orders = snap.get("orders_by_status") or {}
    financial = snap.get("financial") or {}
    lines = [
        f"المنتجات: {products.get('active', 0)} نشط من {products.get('total', 0)}.",
        f"قيمة المخزون: {products.get('stock_value', 0):,} د.ع.",
        f"طلبات تم الطلب: {orders.get('تم الطلب', 0)}، جاري الشحن: {orders.get('جاري الشحن', 0)}.",
        f"ديون الموردين: {financial.get('supplier_debts', 0):,} د.ع، ذمم شركات النقل إلنا: {financial.get('shipping_receivables', 0):,} د.ع.",
    ]
    if findings.get("accounting_audit"):
        audit = findings["accounting_audit"]
        lines.append(
            f"تدقيق محاسبي: اختلاف مخزون {audit.get('stock_imbalances_count', 0)}، فواتير مختلفة {audit.get('invoice_total_mismatches_count', 0)}، هوامش سالبة {audit.get('negative_margin_items_count', 0)}."
        )
    evidence_text = _local_evidence_text(findings.get("query_evidence") or {})
    if evidence_text:
        lines.append(evidence_text)
    return "\n".join(lines)


def _local_evidence_text(evidence: dict) -> str:
    products = evidence.get("products") or []
    if products:
        lines = ["\nالأدلة المحلية المباشرة:"]
        if evidence.get("scope_note"):
            lines.append(f"- نطاق التحليل: {evidence.get('scope_note')}.")
        for product in products[:3]:
            lines.append(
                f"- المنتج #{product.get('product_id')}: {product.get('name')} | إجمالي المخزون {product.get('total_quantity', 0)} | شراء {product.get('buy_price', 0):,} | بيع {product.get('sale_price', 0):,}."
            )
            for stock in (product.get("branch_stock") or [])[:4]:
                lines.append(
                    f"  - {stock.get('branch_name')}: نظام {stock.get('system_qty', 0)}، محجوز/جاري الشحن {stock.get('reserved_ordered_or_shipping', 0)}، قابل للبيع {stock.get('salable_qty', 0)}."
                )
            orders = product.get("recent_orders") or []
            if orders:
                order_bits = [
                    f"#{o.get('invoice_id')} {o.get('status')} {o.get('branch_name')} كمية {o.get('quantity')}"
                    for o in orders[:5]
                ]
                lines.append("  - آخر الطلبات: " + " | ".join(order_bits))
            movements = product.get("recent_movements") or []
            if movements:
                move_bits = [
                    f"{m.get('date')} {m.get('type')} +{m.get('in', 0)}/-{m.get('out', 0)} رصيد {m.get('balance_after', 0)} ({m.get('reference')})"
                    for m in movements[-5:]
                    if not m.get("error")
                ]
                if move_bits:
                    lines.append("  - آخر الحركات: " + " | ".join(move_bits))
        return "\n".join(lines)
    low_stock = evidence.get("low_stock_samples") or []
    if low_stock:
        lines = ["\nعينات منخفضة المخزون من النظام:"]
        for row in low_stock[:8]:
            lines.append(
                f"- {row.get('product_name')} في {row.get('branch_name')}: {row.get('qty')} قطعة، حد التنبيه {row.get('threshold')}."
            )
        return "\n".join(lines)
    return ""


def approve_action_plan(plan_id: int, *, employee_id: int, note: str | None = None) -> AIActionPlan:
    plan = AIActionPlan.query.get_or_404(plan_id)
    if plan.status not in {"draft", "rejected"}:
        raise ValueError("هذه الخطة ليست بانتظار الموافقة")
    plan.status = "approved"
    plan.approved_by_id = employee_id
    plan.approved_at = datetime.utcnow()
    plan.approval_note = note
    _log_tool("action_plan.approve", plan_id=plan.id, employee_id=employee_id, mode="approve")
    db.session.commit()
    return plan


def reject_action_plan(plan_id: int, *, employee_id: int, reason: str | None = None) -> AIActionPlan:
    plan = AIActionPlan.query.get_or_404(plan_id)
    if plan.status == "executed":
        raise ValueError("لا يمكن رفض خطة منفذة")
    plan.status = "rejected"
    plan.rejection_reason = reason or "رفض من الأدمن"
    _log_tool("action_plan.reject", plan_id=plan.id, employee_id=employee_id, mode="reject", input_data={"reason": reason or ""})
    db.session.commit()
    return plan


def validate_action_plan(plan_id: int) -> dict:
    plan = AIActionPlan.query.get_or_404(plan_id)
    errors: list[str] = []
    warnings: list[str] = []
    item_results: list[dict] = []
    for item in plan.items:
        if item.status == "executed":
            item_results.append({"item_id": item.id, "status": "skipped", "message": "منفذ مسبقاً"})
            continue
        result = _preflight_action_item(item)
        item_results.append({"item_id": item.id, **result})
        errors.extend(result.get("errors") or [])
        warnings.extend(result.get("warnings") or [])
    return {
        "ok": not errors,
        "errors": errors,
        "warnings": warnings,
        "items": item_results,
    }


def execute_action_plan(plan_id: int, *, employee_id: int) -> AIActionPlan:
    plan = AIActionPlan.query.get_or_404(plan_id)
    if plan.status != "approved":
        raise ValueError("لا يمكن التنفيذ قبل موافقة الأدمن")
    validation = validate_action_plan(plan_id)
    if not validation["ok"]:
        _log_tool(
            "action_plan.preflight_failed",
            plan_id=plan.id,
            employee_id=employee_id,
            mode="validate",
            output_data=validation,
            status="error",
            error="; ".join(validation["errors"][:5]),
        )
        raise ValueError("فشل فحص الخطة قبل التنفيذ: " + " | ".join(validation["errors"][:5]))
    results: list[dict] = []
    try:
        for item in plan.items:
            if item.status == "executed":
                continue
            result = _execute_action_item(item, employee_id=employee_id)
            item.status = "executed"
            item.executed_at = datetime.utcnow()
            item.set_result(result)
            results.append({"item_id": item.id, **result})
        plan.status = "executed"
        plan.executed_by_id = employee_id
        plan.executed_at = datetime.utcnow()
        plan.set_execution_result({"items": results})
        log_activity(
            "execute",
            "ai_assistant",
            f"تنفيذ خطة مساعد Finora #{plan.id}: {plan.title}",
            entity_type="ai_action_plan",
            entity_id=plan.id,
            payload=plan.to_dict(),
            commit=False,
        )
        _log_tool("action_plan.execute", plan_id=plan.id, employee_id=employee_id, mode="execute", output_data={"items": len(results)})
        db.session.commit()
        return plan
    except Exception:
        db.session.rollback()
        raise


def _preflight_action_item(item: AIActionItem) -> dict:
    payload = item.get_payload()
    before = item.get_before()
    errors: list[str] = []
    warnings: list[str] = []

    def changed(label: str, current, expected) -> None:
        if str(current) != str(expected):
            errors.append(f"{item.title}: تغير {label} من {expected} إلى {current} بعد إنشاء الخطة")

    if item.item_type == "inventory_transfer":
        from_branch_id = int(payload.get("from_branch_id") or 0)
        to_branch_id = int(payload.get("to_branch_id") or 0)
        product_id = int(payload.get("product_id") or 0)
        qty = int(payload.get("quantity") or 0)
        if qty <= 0:
            errors.append(f"{item.title}: كمية التحويل غير صالحة")
        from_stock = get_branch_stock(from_branch_id, product_id)
        to_stock = get_branch_stock(to_branch_id, product_id)
        if "from_branch_stock" in before:
            changed("مخزون الفرع المصدر", from_stock, before.get("from_branch_stock"))
        if "to_branch_stock" in before:
            changed("مخزون الفرع الهدف", to_stock, before.get("to_branch_stock"))
        if from_stock < qty:
            errors.append(f"{item.title}: المخزون الحالي في الفرع المصدر ({from_stock}) أقل من كمية التحويل ({qty})")

    elif item.item_type == "stock_adjustment":
        branch_id = int(payload.get("branch_id") or 0)
        product_id = int(payload.get("product_id") or 0)
        adjustment = int(payload.get("adjustment") or 0)
        current = get_branch_stock(branch_id, product_id)
        if "branch_stock" in before:
            changed("مخزون الفرع", current, before.get("branch_stock"))
        if current + adjustment < 0:
            errors.append(f"{item.title}: التسوية تجعل المخزون سالباً ({current + adjustment})")

    elif item.item_type == "order_update":
        invoice = Invoice.query.get(int(payload.get("invoice_id") or 0))
        if not invoice:
            errors.append(f"{item.title}: الطلب غير موجود")
        else:
            changed("حالة الطلب", invoice.status, before.get("status"))
            changed("حالة الدفع", invoice.payment_status, before.get("payment_status"))
            changed("المبلغ المسدد", int(invoice.paid_amount or 0), int(before.get("paid_amount") or 0))
            action = payload.get("action")
            if action == "cancel" and (invoice.status or "").strip() != "تم الطلب":
                errors.append(f"{item.title}: الإلغاء مسموح فقط للطلبات بحالة تم الطلب")
            if action == "return" and not (payload.get("barcode") or "").strip():
                errors.append(f"{item.title}: الإرجاع يحتاج باركود داخل الخطة")
            if action == "mark_paid" and (invoice.status or "").strip() in {"ملغي", "مرتجع", "راجع", "راجعة"}:
                errors.append(f"{item.title}: لا يمكن تسديد طلب ملغي أو راجع")

    elif item.item_type == "supplier_ledger_fix":
        supplier = Supplier.query.get(int(payload.get("supplier_id") or 0))
        if not supplier:
            errors.append(f"{item.title}: المورد غير موجود")
        else:
            changed("دين المورد", int(supplier.total_debt or 0), int(before.get("total_debt") or 0))
            changed("مدفوع المورد", int(supplier.total_paid or 0), int(before.get("total_paid") or 0))

    elif item.item_type == "shipping_opening_balance_fix":
        company = ShippingCompany.query.get(int(payload.get("shipping_company_id") or 0))
        if not company:
            errors.append(f"{item.title}: شركة النقل غير موجودة")
        else:
            changed("الرصيد الافتتاحي", int(company.opening_balance or 0), int(before.get("opening_balance") or 0))

    elif item.item_type == "shipping_payment_delete_duplicate":
        payment = ShippingPayment.query.get(int(payload.get("payment_id") or 0))
        if not payment:
            errors.append(f"{item.title}: حركة القبض غير موجودة أو حذفت مسبقاً")
        else:
            changed("شركة النقل", payment.shipping_company_id, before.get("shipping_company_id"))
            changed("رقم الطلب", payment.invoice_id, before.get("invoice_id"))
            changed("المبلغ", int(payment.amount or 0), int(before.get("amount") or 0))
            changed("نوع الحركة", payment.action or "", before.get("action") or "")
            duplicate_ids = [int(value) for value in payload.get("duplicate_payment_ids") or []]
            matching = ShippingPayment.query.filter(ShippingPayment.id.in_(duplicate_ids)).all()
            exact = [
                row for row in matching
                if row.shipping_company_id == payment.shipping_company_id
                and row.invoice_id == payment.invoice_id
                and int(row.amount or 0) == int(payment.amount or 0)
                and (row.action or "") == (payment.action or "")
                and row.treasury_account_id == payment.treasury_account_id
            ]
            if len(exact) < 2:
                errors.append(f"{item.title}: لم تعد الحركة مكررة؛ تم إيقاف الحذف")
            if payload.get("restore_opening_balance"):
                company = ShippingCompany.query.get(payment.shipping_company_id)
                if not company:
                    errors.append(f"{item.title}: شركة النقل غير موجودة")
                else:
                    changed("الرصيد الافتتاحي", int(company.opening_balance or 0), int(before.get("opening_balance") or 0))

    elif item.item_type == "fixed_asset_set_capital_and_post":
        asset = FixedAsset.query.get(int(payload.get("asset_id") or 0))
        if not asset:
            errors.append(f"{item.title}: الأصل غير موجود")
        else:
            changed("حالة الأصل", asset.status, before.get("status"))
            changed("طريقة الدفع", asset.payment_method, before.get("payment_method"))
            changed("المدفوع", int(asset.paid_amount or 0), int(before.get("paid_amount") or 0))
            changed("الآجل", int(asset.credit_amount or 0), int(before.get("credit_amount") or 0))
            changed("التكلفة", int(asset.total_cost or 0), int(before.get("total_cost") or 0))
            if asset.acquisition_journal_entry_id:
                errors.append(f"{item.title}: الأصل مرحّل مسبقاً ولا يمكن تغيير تمويله بهذه الخطة")
            if int(asset.total_cost or 0) <= 0:
                errors.append(f"{item.title}: تكلفة الأصل غير صالحة")

    elif item.item_type == "fixed_asset_create_and_post":
        category = FixedAssetCategory.query.get(int(payload.get("category_id") or 0))
        if not category or not category.is_active:
            errors.append(f"{item.title}: تصنيف الأصل غير موجود أو غير نشط")
        amount = int(payload.get("purchase_price") or 0)
        if amount <= 0:
            errors.append(f"{item.title}: مبلغ شراء الأصل غير صالح")
        purchase_date = datetime.strptime(payload.get("purchase_date"), "%Y-%m-%d").date()
        duplicates = FixedAsset.query.filter(
            FixedAsset.name == payload.get("name"),
            FixedAsset.purchase_price == amount,
            FixedAsset.purchase_date == purchase_date,
        ).count()
        if duplicates:
            errors.append(f"{item.title}: يوجد أصل مطابق بالاسم والمبلغ والتاريخ؛ أوقفنا التكرار")
        if payload.get("payment_method") == "cash":
            cash = get_default_cash_account()
            balance = int(calculate_treasury_balance(cash.id) or 0)
            if balance < amount:
                errors.append(f"{item.title}: رصيد الصندوق {balance:,} أقل من مبلغ الأصل {amount:,} د.ع")

    elif item.item_type == "review_task":
        warnings.append(f"{item.title}: مهمة مراجعة فقط ولا تعدل بيانات")

    else:
        errors.append(f"{item.title}: نوع إجراء غير مدعوم ({item.item_type})")

    return {"ok": not errors, "errors": errors, "warnings": warnings}


def _execute_action_item(item: AIActionItem, *, employee_id: int) -> dict:
    payload = item.get_payload()
    if item.item_type == "inventory_transfer":
        from_branch_id = int(payload["from_branch_id"])
        to_branch_id = int(payload["to_branch_id"])
        product_id = int(payload["product_id"])
        qty = int(payload["quantity"])
        if qty <= 0:
            raise ValueError("كمية التحويل غير صالحة")
        transfer_deduct(from_branch_id, [(product_id, qty)], sync_product=False)
        transfer_receive(to_branch_id, [(product_id, qty)], sync_product=False)
        sync_product_total(product_id)
        transfer = StockTransfer(
            transfer_no=f"AI-TR-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{item.id}",
            from_branch_id=from_branch_id,
            to_branch_id=to_branch_id,
            status="received",
            note=payload.get("reason") or "تحويل مقترح من مساعد Finora",
            created_by_id=employee_id,
            received_by_id=employee_id,
            sent_at=datetime.utcnow(),
            received_at=datetime.utcnow(),
        )
        db.session.add(transfer)
        db.session.flush()
        db.session.add(
            StockTransferLine(
                transfer_id=transfer.id,
                product_id=product_id,
                quantity=qty,
                quantity_received=qty,
            )
        )
        return {"transfer_id": transfer.id, "product_id": product_id, "quantity": qty}

    if item.item_type == "stock_adjustment":
        branch_id = int(payload["branch_id"])
        product_id = int(payload["product_id"])
        adjustment = int(payload["adjustment"])
        before = get_branch_stock(branch_id, product_id)
        adjust_branch_stock(branch_id, product_id, adjustment, sync_product=True)
        after = get_branch_stock(branch_id, product_id)
        return {"branch_id": branch_id, "product_id": product_id, "before": before, "after": after, "adjustment": adjustment}

    if item.item_type == "order_update":
        invoice = Invoice.query.get(int(payload["invoice_id"]))
        if not invoice:
            raise ValueError("الطلب غير موجود")
        before = {"status": invoice.status, "payment_status": invoice.payment_status, "paid_amount": invoice.paid_amount}
        action = payload.get("action")
        prev_effective_paid = _effective_paid_amount(invoice)
        if action == "mark_paid":
            ensure_stock_for_transition(invoice, target_status="تم التوصيل", target_payment_status="مسدد")
            invoice.payment_status = "مسدد"
            invoice.paid_amount = int(invoice.total or 0)
            if invoice.status not in {"تم التوصيل", "مرتجع", "راجع", "راجعة"}:
                invoice.status = "تم التوصيل"
        elif action == "cancel":
            process_order_cancel(invoice)
        elif action == "return":
            barcode = (payload.get("barcode") or "").strip()
            if not barcode:
                raise ValueError("إرجاع الطلب يحتاج باركود الطلب أو باركود شركة النقل داخل الخطة")
            process_order_return(invoice, barcode)
        else:
            raise ValueError("نوع تعديل الطلب غير مدعوم")
        delta_pay = _effective_paid_amount(invoice) - prev_effective_paid
        append_payment_ledger_delta(invoice.id, delta_pay)
        sync_delivery_expense_for_invoice(invoice)
        after = {"status": invoice.status, "payment_status": invoice.payment_status, "paid_amount": invoice.paid_amount}
        return {"invoice_id": invoice.id, "before": before, "after": after}

    if item.item_type == "supplier_ledger_fix":
        supplier = Supplier.query.get(int(payload["supplier_id"]))
        if not supplier:
            raise ValueError("المورد غير موجود")
        before = {
            "total_debt": int(supplier.total_debt or 0),
            "total_paid": int(supplier.total_paid or 0),
        }
        supplier.total_debt = int(payload["expected_debt"])
        supplier.total_paid = int(payload["expected_paid"])
        after = {
            "total_debt": int(supplier.total_debt or 0),
            "total_paid": int(supplier.total_paid or 0),
        }
        return {"supplier_id": supplier.id, "before": before, "after": after}

    if item.item_type == "shipping_opening_balance_fix":
        company = ShippingCompany.query.get(int(payload["shipping_company_id"]))
        if not company:
            raise ValueError("شركة النقل غير موجودة")
        before = {"opening_balance": int(company.opening_balance or 0)}
        company.opening_balance = int(payload["opening_balance"])
        after = {"opening_balance": int(company.opening_balance or 0)}
        return {"shipping_company_id": company.id, "before": before, "after": after}

    if item.item_type == "shipping_payment_delete_duplicate":
        payment = ShippingPayment.query.get(int(payload["payment_id"]))
        if not payment:
            raise ValueError("حركة القبض غير موجودة")
        company = ShippingCompany.query.get(payment.shipping_company_id)
        before = {
            "payment_id": payment.id,
            "amount": int(payment.amount or 0),
            "opening_balance": int(company.opening_balance or 0) if company else None,
        }
        if payload.get("restore_opening_balance"):
            if not company:
                raise ValueError("شركة النقل غير موجودة")
            company.opening_balance = int(company.opening_balance or 0) + int(payment.amount or 0)
        deleted_id = payment.id
        amount = int(payment.amount or 0)
        db.session.delete(payment)
        db.session.flush()
        still_exists = ShippingPayment.query.get(deleted_id) is not None
        return {
            "payment_id": deleted_id,
            "deleted": not still_exists,
            "amount": amount,
            "before": before,
            "after": {
                "payment_exists": still_exists,
                "opening_balance": int(company.opening_balance or 0) if company else None,
            },
        }

    if item.item_type == "fixed_asset_set_capital_and_post":
        asset = FixedAsset.query.get(int(payload["asset_id"]))
        if not asset:
            raise ValueError("الأصل غير موجود")
        cash = get_default_cash_account()
        cash_before = int(calculate_treasury_balance(cash.id) or 0)
        before = {
            "status": asset.status,
            "payment_method": asset.payment_method,
            "paid_amount": int(asset.paid_amount or 0),
            "credit_amount": int(asset.credit_amount or 0),
        }
        asset.payment_method = "capital"
        asset.treasury_account_id = None
        asset.paid_amount = 0
        asset.credit_amount = 0
        post_asset_acquisition(asset, user_id=employee_id)
        db.session.flush()
        cash_after = int(calculate_treasury_balance(cash.id) or 0)
        return {
            "asset_id": asset.id,
            "asset_code": asset.asset_code,
            "before": before,
            "after": {
                "status": asset.status,
                "payment_method": asset.payment_method,
                "paid_amount": int(asset.paid_amount or 0),
                "credit_amount": int(asset.credit_amount or 0),
                "posted": bool(asset.acquisition_journal_entry_id),
                "cash_before": cash_before,
                "cash_after": cash_after,
            },
        }

    if item.item_type == "fixed_asset_create_and_post":
        payment_method = payload.get("payment_method") or "credit"
        cash = get_default_cash_account()
        cash_before = int(calculate_treasury_balance(cash.id) or 0)
        form_data = {
            "name": payload["name"],
            "category_id": payload["category_id"],
            "purchase_price": payload["purchase_price"],
            "purchase_date": payload["purchase_date"],
            "payment_method": payment_method,
            "treasury_account_id": cash.id if payment_method == "cash" else None,
        }
        asset = build_asset_from_form(form_data, user_id=employee_id, as_draft=True)
        post_asset_acquisition(asset, user_id=employee_id)
        db.session.flush()
        cash_after = int(calculate_treasury_balance(cash.id) or 0)
        return {
            "asset_id": asset.id,
            "asset_code": asset.asset_code,
            "after": {
                "name": asset.name,
                "status": asset.status,
                "total_cost": int(asset.total_cost or 0),
                "payment_method": asset.payment_method,
                "posted": bool(asset.acquisition_journal_entry_id),
                "cash_before": cash_before,
                "cash_after": cash_after,
                "cash_effect": cash_after - cash_before,
            },
        }

    if item.item_type == "review_task":
        return {"review_only": True, "message": "هذه مهمة مراجعة فقط ولا تعدل بيانات النظام."}

    raise ValueError(f"نوع إجراء غير مدعوم: {item.item_type}")


def run_ai_audit(
    *,
    audit_type: str = "comprehensive",
    schedule_id: int | None = None,
    employee_id: int | None = None,
) -> AIAuditRun:
    run = AIAuditRun(schedule_id=schedule_id, run_type=audit_type, status="running")
    db.session.add(run)
    db.session.flush()
    try:
        result: dict[str, Any] = {"snapshot": collect_system_snapshot()}
        result["financial_labels"] = {
            "supplier_debts": "التزامات على الشركة للموردين",
            "shipping_receivables": "ذمم مدينة إلنا عند شركات النقل",
        }
        if audit_type in {"accounting", "comprehensive", "inventory", "orders"}:
            result["accounting"] = audit_accounting_integrity(limit=200)
        issue_count = 0
        if result.get("accounting"):
            summary = result["accounting"].get("summary") or {}
            issue_count = sum(
                int(summary.get(key, 0) or 0)
                for key in (
                    "stock_imbalances_count",
                    "status_inconsistencies_count",
                    "invoice_total_mismatches_count",
                    "negative_margin_items_count",
                )
            )
            if issue_count:
                review_plan = _build_audit_review_plan(
                    audit_type=audit_type,
                    summary=summary,
                    employee_id=employee_id,
                    schedule_id=schedule_id,
                )
                run.action_plan_id = review_plan.id
        severity = "critical" if issue_count >= 10 else ("warning" if issue_count else "success")
        title = "تدقيق مساعد Finora"
        summary_text = f"اكتشف التدقيق {issue_count} ملاحظة." if issue_count else "لم يجد التدقيق أخطاء حرجة."
        analytics = SystemAnalytics(
            analysis_type=f"ai_{audit_type}",
            title=title,
            description=summary_text,
            severity=severity,
            affected_count=issue_count,
            related_data=_json_dumps(result),
        )
        db.session.add(analytics)
        schedule = AIScheduledAudit.query.get(schedule_id) if schedule_id else None
        alert_created = False
        if _severity_reaches_threshold(severity, schedule.severity_threshold if schedule else "warning"):
            alert_created = _add_ai_alert_once(
                title=title,
                message=summary_text,
                severity=severity,
                related_id=run.id,
            )
        if severity == "critical" and alert_created:
            _notify_admins_internal(f"{title}: {summary_text}")
        run.status = "completed"
        run.summary = summary_text
        run.set_result(result)
        run.finished_at = datetime.utcnow()
        if schedule:
            schedule.last_run_at = datetime.utcnow()
            schedule.next_run_at = datetime.utcnow() + timedelta(minutes=int(schedule.interval_minutes or 1440))
        _log_tool("scheduled_audit.run", employee_id=employee_id, input_data={"audit_type": audit_type}, output_data={"issues": issue_count})
        db.session.commit()
        return run
    except Exception as exc:
        db.session.rollback()
        run = AIAuditRun.query.get(run.id) if run.id else run
        if run:
            run.status = "failed"
            run.summary = str(exc)
            run.finished_at = datetime.utcnow()
            db.session.add(run)
            db.session.commit()
        raise


def list_schedules() -> list[dict]:
    return [s.to_dict() for s in AIScheduledAudit.query.order_by(AIScheduledAudit.created_at.desc()).all()]


def create_or_update_schedule(data: dict, *, employee_id: int | None) -> AIScheduledAudit:
    raw_schedule_id = data.get("id")
    try:
        schedule_id = int(raw_schedule_id) if raw_schedule_id not in (None, "") else None
    except (TypeError, ValueError):
        schedule_id = None
    schedule = AIScheduledAudit.query.get(schedule_id) if schedule_id else None
    if not schedule:
        schedule = AIScheduledAudit(created_by_id=employee_id)
        db.session.add(schedule)
    schedule.name = (data.get("name") or "تدقيق دوري").strip()[:180]
    audit_type = (data.get("audit_type") or "comprehensive").strip()
    if audit_type not in {"comprehensive", "accounting", "inventory", "orders", "suppliers", "shipping"}:
        audit_type = "comprehensive"
    schedule.audit_type = audit_type
    try:
        interval_minutes = int(data.get("interval_minutes") or 1440)
    except (TypeError, ValueError):
        interval_minutes = 1440
    schedule.interval_minutes = max(15, interval_minutes)
    severity_threshold = (data.get("severity_threshold") or "warning").strip()
    if severity_threshold not in {"info", "warning", "critical"}:
        severity_threshold = "warning"
    schedule.severity_threshold = severity_threshold
    schedule.is_active = bool(data.get("is_active", True))
    schedule.set_settings(data.get("settings") or {})
    if not schedule.next_run_at:
        schedule.next_run_at = datetime.utcnow() + timedelta(minutes=schedule.interval_minutes)
    db.session.commit()
    return schedule


def run_due_scheduled_audits(limit: int = 5) -> int:
    now = datetime.utcnow()
    schedules = (
        AIScheduledAudit.query.filter_by(is_active=True)
        .filter((AIScheduledAudit.next_run_at.is_(None)) | (AIScheduledAudit.next_run_at <= now))
        .order_by(AIScheduledAudit.next_run_at.asc().nullsfirst())
        .limit(limit)
        .all()
    )
    count = 0
    for schedule in schedules:
        try:
            run_ai_audit(audit_type=schedule.audit_type, schedule_id=schedule.id)
            count += 1
        except Exception:
            current_app.logger.exception("AI scheduled audit failed for schedule %s", schedule.id)
    return count


def _build_audit_review_plan(
    *,
    audit_type: str,
    summary: dict,
    employee_id: int | None,
    schedule_id: int | None,
) -> AIActionPlan:
    issue_map = [
        ("stock_imbalances_count", "مراجعة فروقات المخزون", "طابق branch_stock مع product.quantity وحركات المخزون."),
        ("status_inconsistencies_count", "مراجعة حالات الطلبات", "راجع الطلبات ذات الحالة أو الدفع غير المتناسق."),
        ("invoice_total_mismatches_count", "مراجعة إجماليات الفواتير", "راجع فواتير يختلف مجموع عناصرها عن الإجمالي المحفوظ."),
        ("negative_margin_items_count", "مراجعة الهوامش السالبة", "راجع أسعار الشراء/البيع للمنتجات التي تظهر بخسارة."),
    ]
    total_issues = sum(int(summary.get(key, 0) or 0) for key, _, _ in issue_map)
    plan = AIActionPlan(
        created_by_id=employee_id,
        title=f"خطة مراجعة تدقيق {audit_type}",
        plan_type="scheduled_audit_review",
        status="draft",
        summary="خطة مراجعة مقترحة من الدورية. لا تنفذ أي تعديل تلقائياً.",
        risk_level="high" if total_issues >= 10 else "medium",
    )
    plan.set_impact(
        {
            "audit_type": audit_type,
            "schedule_id": schedule_id,
            "summary": summary,
            "note": "هذه خطة مراجعة؛ أي إصلاح فعلي يحتاج خطة تنفيذ منفصلة وموافقة أدمن.",
        }
    )
    db.session.add(plan)
    db.session.flush()
    created_items = 0
    for key, title, description in issue_map:
        count = int(summary.get(key, 0) or 0)
        if not count:
            continue
        item = AIActionItem(
            plan_id=plan.id,
            item_type="review_task",
            target_type="ai_audit",
            target_id=None,
            title=f"{title}: {count}",
            description=description,
        )
        item.set_before({"count": count, "source": key})
        item.set_after({"expected": "مراجعة بشرية أو إنشاء خطة إصلاح منفصلة"})
        item.set_payload({"audit_key": key, "audit_type": audit_type})
        db.session.add(item)
        created_items += 1
    _log_tool(
        "scheduled_audit.build_review_plan",
        plan_id=plan.id,
        employee_id=employee_id,
        input_data={"audit_type": audit_type},
        output_data={"items": created_items, "summary": summary},
        mode="plan",
    )
    return plan


def _severity_reaches_threshold(severity: str, threshold: str | None) -> bool:
    order = {"success": 0, "info": 1, "warning": 2, "critical": 3}
    severity_value = order.get((severity or "info").lower(), 1)
    threshold_value = order.get((threshold or "warning").lower(), 2)
    return severity_value >= threshold_value and severity_value > 0


def _add_ai_alert_once(*, title: str, message: str, severity: str, related_id: int | None) -> bool:
    if severity not in {"critical", "warning", "info"}:
        return False
    recent_cutoff = datetime.utcnow() - timedelta(hours=6)
    existing = (
        SystemAlert.query.filter_by(alert_type="ai_audit", title=title, message=message, is_dismissed=False)
        .filter(SystemAlert.created_at >= recent_cutoff)
        .first()
    )
    if existing:
        return False
    db.session.add(
        SystemAlert(
            alert_type="ai_audit",
            title=title,
            message=message,
            priority="high" if severity == "critical" else ("medium" if severity == "warning" else "low"),
            related_type="ai_audit_run",
            related_id=related_id,
        )
    )
    return True


def _notify_admins_internal(content: str) -> None:
    admins = Employee.query.filter_by(role="admin", is_active=True).limit(10).all()
    for admin in admins:
        db.session.add(
            Message(
                sender_id=admin.id,
                receiver_id=admin.id,
                content=content,
                file_type=None,
                is_read=False,
            )
        )
