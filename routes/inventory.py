import json
import os
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, current_app, g, send_file
from werkzeug.utils import secure_filename
from extensions import db
from models.product import Product
from models.order_item import OrderItem
from models.invoice import Invoice
from models.supplier import Supplier
# من models.purchase تم نقله إلى purchases.py
from models.employee import Employee
from models.account_transaction import AccountTransaction
from sqlalchemy.exc import IntegrityError, SQLAlchemyError
from sqlalchemy.sql import func
from datetime import datetime

# دوال حركات المخزون (للعرض فقط)
from utils.inventory_movements import (
    get_product_inventory_movements,
    get_product_inventory_summary,
    get_low_stock_products,
    get_out_of_stock_products,
    validate_sale_quantity
)
from utils.permission_checks import check_permission
from utils.activity_logger import PRODUCT_SNAPSHOT_FIELDS, log_activity, log_mutation, snapshot_attrs

from utils.product_schema_guard import ensure_product_schema
from utils.order_item_schema_guard import ensure_order_item_schema
from utils.order_shipping import sync_product_name_to_order_items
from utils.product_delivery_fees import (
    apply_delivery_fees_to_meta,
    delivery_fees_from_form,
    product_delivery_config,
)
from utils.tenant_provinces import get_tenant_provinces_config
from utils.branch_migration import ensure_branch_schema, get_default_branch
from utils.branch_context import current_branch_id, init_branch_context
from utils.branch_stock_service import (
    adjust_branch_stock,
    branch_stock_map,
    get_branch_stock,
    set_branch_stock,
    set_opening_branch_stock,
    BranchStockError,
)

inventory_bp = Blueprint("inventory", __name__)


@inventory_bp.before_request
def _inventory_branch_setup():
    if "user_id" not in session or not getattr(g, "tenant", None):
        return
    ensure_product_schema()
    ensure_order_item_schema()
    ensure_branch_schema()
    init_branch_context()


def _inventory_branch_id():
    if getattr(g, "view_all_branches", False):
        default = get_default_branch()
        return default.id if default else None
    return current_branch_id() or (get_default_branch().id if get_default_branch() else None)


def _branch_id_for_product_stock(form=None, meta: dict | None = None) -> int | None:
    """Use the product's selected branch for stock; fallback to current inventory branch."""
    from models.branch import Branch

    candidates: list[int] = []
    if form is not None:
        raw = (form.get("branch_id") or "").strip()
        if raw.isdigit():
            candidates.append(int(raw))
    if meta is not None:
        raw = meta.get("branch_id")
        if raw is not None:
            try:
                candidates.append(int(raw))
            except (TypeError, ValueError):
                pass

    seen: set[int] = set()
    for branch_id in candidates:
        if branch_id <= 0 or branch_id in seen:
            continue
        seen.add(branch_id)
        branch = Branch.query.filter_by(id=branch_id, is_active=True).first()
        if branch:
            return branch.id

    return _inventory_branch_id()


def _edit_current_stock_for_product(product: Product, meta: dict | None = None) -> int:
    """Current on-hand quantity for the product edit form (branch-scoped when applicable)."""
    stock_branch_id = _branch_id_for_product_stock(meta=meta)
    if stock_branch_id:
        return get_branch_stock(stock_branch_id, product.id)
    return int(product.quantity or 0)


def _product_display_qty(product, stock_map, view_all):
    fallback = int(product.quantity or 0)
    return stock_map.get(product.id, fallback)


def _load_product_meta(product) -> dict:
    raw = ((product.meta_json if product else None) or "").strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _split_multiline_values(raw) -> list[str]:
    items: list[str] = []
    for line in str(raw or "").splitlines():
        value = line.strip()
        if value:
            items.append(value)
    return items


def _parse_specs_text(raw) -> list[dict]:
    items: list[dict] = []
    for line in str(raw or "").splitlines():
        row = line.strip(" -\t\r\n")
        if not row:
            continue
        if ":" in row:
            label, value = row.split(":", 1)
        elif " - " in row:
            label, value = row.split(" - ", 1)
        else:
            label, value = "تفصيل", row
        value = value.strip()
        label = label.strip() or "تفصيل"
        if value:
            items.append({"label": label, "value": value})
    return items


def _parse_optional_date(raw):
    raw = (raw or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None


def _extract_specs_items(meta: dict | None) -> list[dict]:
    meta = meta or {}
    raw_items = meta.get("specs_items")
    items: list[dict] = []
    if isinstance(raw_items, list):
        for row in raw_items:
            if not isinstance(row, dict):
                continue
            label = str(row.get("label") or "").strip() or "تفصيل"
            value = str(row.get("value") or "").strip()
            if value:
                items.append({"label": label, "value": value})
    if items:
        return items
    return _parse_specs_text(meta.get("specs_text"))


def _specs_items_from_form(form) -> list[dict]:
    labels = form.getlist("spec_label[]")
    values = form.getlist("spec_value[]")
    items: list[dict] = []
    count = max(len(labels), len(values))
    for idx in range(count):
        label = str(labels[idx] if idx < len(labels) else "").strip() or "تفصيل"
        value = str(values[idx] if idx < len(values) else "").strip()
        if value:
            items.append({"label": label, "value": value})
    if items:
        return items
    return _parse_specs_text(form.get("specs_text"))


def _meta_from_inventory_add_form(form) -> dict:
    """بناء meta_json من نموذج صفحة إضافة/تعديل المنتج."""
    meta_keys = (
        "barcode_type",
        "unit",
        "brand",
        "category",
        "subcategory",
        "warranty",
        "tax_applied",
        "sales_tax_type",
        "product_type",
        "shelf",
        "shelf_row",
        "shelf_loc",
        "weight",
        "custom_field_1",
        "custom_field_2",
        "custom_field_3",
        "custom_field_4",
        "purchase_ex_tax",
        "purchase_inc_tax",
        "sale_ex_tax",
        "sale_inc_tax",
        "beauty_product_type",
        "usage_instructions",
        "warning_notes",
        "shelf_life_after_opening_days",
    )
    meta: dict = {}
    for k in meta_keys:
        v = (form.get(k) or "").strip()
        if v:
            meta[k] = v

    branch_id_raw = (form.get("branch_id") or "").strip()
    if branch_id_raw.isdigit():
        from models.branch import Branch

        branch_id = int(branch_id_raw)
        branch = Branch.query.filter_by(id=branch_id, is_active=True).first()
        if branch:
            meta["branch_id"] = branch.id
            meta["branch_name"] = branch.name
            meta["branch_note"] = branch.name
    else:
        meta.pop("branch_id", None)
        meta.pop("branch_name", None)
        meta.pop("branch_note", None)

    if form.get("enable_imei"):
        meta["enable_imei"] = True
    if bool(form.get("not_for_sale")):
        meta["not_for_sale"] = True
    video_url = (form.get("video_url") or "").strip()
    if video_url:
        meta["video_url"] = video_url
    gallery_urls = _split_multiline_values(form.get("gallery_urls"))
    if gallery_urls:
        meta["gallery"] = gallery_urls
    specs_items = _specs_items_from_form(form)
    if specs_items:
        meta["specs_items"] = specs_items
    store_badge = (form.get("store_badge") or "").strip()
    if store_badge:
        meta["store_badge"] = store_badge
    by_province, default_fee = delivery_fees_from_form(
        form,
        get_tenant_provinces_config().get("list") or [],
    )
    meta = apply_delivery_fees_to_meta(meta, by_province, default_fee)
    if form.get("has_colors"):
        meta["has_colors"] = True
    return meta


def _color_rows_from_form(form) -> list[tuple[str, int]]:
    names = form.getlist("color_name[]")
    qtys = form.getlist("color_qty[]")
    rows: list[tuple[str, int]] = []
    for name, qty in zip(names, qtys):
        color = (name or "").strip()
        if not color:
            continue
        rows.append((color, max(0, int(qty or 0))))
    return rows


def _apply_product_colors_from_form(product: Product, form) -> int:
    from utils.product_color_service import save_product_colors
    from models.product_color_variant import ProductColorVariant

    has_colors = bool(form.get("has_colors"))
    if not has_colors:
        ProductColorVariant.query.filter_by(product_id=product.id).delete(synchronize_session=False)
        meta = _load_product_meta(product)
        meta.pop("has_colors", None)
        product.meta_json = json.dumps(meta, ensure_ascii=False) if meta else None
        return int(product.opening_stock or product.quantity or 0)

    rows = _color_rows_from_form(form)
    if not rows:
        return 0
    save_product_colors(product.id, rows)
    return sum(qty for _, qty in rows)


def _company_branches_for_form():
    ensure_branch_schema()
    from models.branch import Branch

    return (
        Branch.query.filter_by(is_active=True)
        .order_by(Branch.is_default.desc(), Branch.name.asc())
        .all()
    )


def _inventory_add_summary():
    """إحصائيات مختصرة لشريط الملخص في صفحة إضافة المنتج."""
    products = Product.query.all()

    def _val(v, default=0):
        return default if v is None else v

    current_inventory_value = sum(_val(p.buy_price) * _val(p.quantity) for p in products)
    expected_profit_from_stock = sum(
        (
            _val(p.sale_price)
            - _val(p.buy_price)
            - _val(p.shipping_cost)
            - _val(p.marketing_cost)
        )
        * _val(p.quantity)
        for p in products
    )
    return {
        "products_count": len(products),
        "current_inventory_value": current_inventory_value,
        "expected_profit_from_stock": expected_profit_from_stock,
        "company_branches": _company_branches_for_form(),
    }


def _product_branch_info(product, branch_by_id: dict | None = None) -> dict:
    meta = _load_product_meta(product)
    branch_id = meta.get("branch_id")
    label = (meta.get("branch_name") or meta.get("branch_note") or "").strip()
    parsed_id = None
    if branch_id is not None and str(branch_id).strip() != "":
        try:
            parsed_id = int(branch_id)
        except (TypeError, ValueError):
            parsed_id = None
    if parsed_id and branch_by_id:
        branch = branch_by_id.get(parsed_id)
        if branch:
            label = branch.name
    if not label:
        label = "—"
    return {"branch_id": parsed_id, "branch_name": label}


def _product_branch_map(products, branches) -> dict[int, dict]:
    branch_by_id = {b.id: b for b in (branches or [])}
    return {p.id: _product_branch_info(p, branch_by_id) for p in products}


def _product_branch_stock_map(products) -> dict[int, dict[int, int]]:
    from models.branch import BranchStock

    product_ids = [p.id for p in (products or [])]
    if not product_ids:
        return {}

    rows = BranchStock.query.filter(BranchStock.product_id.in_(product_ids)).all()
    stock_by_product: dict[int, dict[int, int]] = {pid: {} for pid in product_ids}
    for row in rows:
        stock_by_product.setdefault(row.product_id, {})[row.branch_id] = int(row.quantity or 0)
    return stock_by_product


def _delivery_fees_context(product=None, meta: dict | None = None) -> dict:
    def _fee_int(value) -> int:
        try:
            return max(0, int(value or 0))
        except (TypeError, ValueError):
            return 0

    cfg = product_delivery_config(product) if product else {"delivery_by_province": {}, "delivery_default_fee": 0}
    if meta:
        by_province = meta.get("delivery_by_province")
        if isinstance(by_province, dict):
            cfg["delivery_by_province"] = {
                str(k).strip(): _fee_int(v)
                for k, v in by_province.items()
                if str(k).strip()
            }
        if meta.get("delivery_default_fee") is not None:
            cfg["delivery_default_fee"] = _fee_int(meta.get("delivery_default_fee"))
    return {
        "delivery_by_province": cfg.get("delivery_by_province") or {},
        "delivery_default_fee": cfg.get("delivery_default_fee") or 0,
    }


def _safe_xlsx_filename(value: str) -> str:
    clean = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(value or "").strip())
    return clean.strip("_") or "inventory"


def _is_shipping_fee_product(product) -> bool:
    marker_values = (product.name or "", product.sku or "", product.barcode or "")
    return any("رسوم الشحن" in str(v) or "__SF_SHIPPING__" in str(v) for v in marker_values)


def _build_inventory_audit_workbook(branch=None, *, all_branches=False):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from models.branch import Branch, BranchStock

    ensure_branch_schema()
    branches = (
        Branch.query.filter_by(is_active=True).order_by(Branch.name.asc()).all()
        if all_branches
        else ([branch] if branch else [])
    )
    if not branches:
        branches = Branch.query.filter_by(is_active=True).order_by(Branch.name.asc()).all()

    products = [
        p for p in Product.query.order_by(Product.name.asc(), Product.id.asc()).all()
        if not _is_shipping_fee_product(p)
    ]
    product_ids = [p.id for p in products]
    branch_ids = [b.id for b in branches]

    stock_by_key: dict[tuple[int, int], int] = {}
    if product_ids and branch_ids:
        rows = (
            BranchStock.query
            .filter(BranchStock.product_id.in_(product_ids), BranchStock.branch_id.in_(branch_ids))
            .all()
        )
        stock_by_key = {
            (int(row.branch_id), int(row.product_id)): int(row.quantity or 0)
            for row in rows
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "الجرد"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A7"
    summary = wb.create_sheet("ملخص الفروع")
    summary.sheet_view.rightToLeft = True
    info = wb.create_sheet("تعليمات")
    info.sheet_view.rightToLeft = True

    title_fill = PatternFill("solid", fgColor="0F172A")
    header_fill = PatternFill("solid", fgColor="2563EB")
    note_fill = PatternFill("solid", fgColor="EFF6FF")
    input_fill = PatternFill("solid", fgColor="FEF9C3")
    light_fill = PatternFill("solid", fgColor="F8FAFC")
    total_fill = PatternFill("solid", fgColor="E0F2FE")
    border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    headers = [
        "رقم الفرع",
        "رقم المنتج",
        "الفرع",
        "المنتج",
        "SKU",
        "الباركود",
        "الحالة",
        "كمية النظام بالفرع",
        "الكمية الفعلية",
        "الفرق",
        "ملاحظة الجرد",
        "إجمالي المنتج بالنظام",
        "سعر الشراء",
    ]
    start_row = 7
    data_rows_count = len(branches) * len(products)
    end_row = start_row + max(data_rows_count, 1) - 1
    scope = "كل الفروع" if all_branches else (branch.name if branch else "كل الفروع")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = "قالب جرد المخزون حسب الفروع"
    ws["A1"].fill = title_fill
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = f"الفرع: {scope} | تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')} | اكتب العدد الفعلي في العمود الأصفر."
    ws["A2"].fill = note_fill
    ws["A2"].font = Font(color="1E3A8A", bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    for cell_range, label in (
        ("A4:B5", "إجمالي أسطر الجرد"),
        ("C4:D5", None),
        ("E4:F5", "الأسطر المدخلة"),
        ("G4:H5", None),
        ("I4:J5", "أسطر بيها فرق"),
        ("K4:M5", None),
    ):
        ws.merge_cells(cell_range)
        if label:
            ws[cell_range.split(":")[0]] = label
    ws["C4"] = f"=COUNTA(D{start_row}:D{end_row})"
    ws["G4"] = f"=COUNT(I{start_row}:I{end_row})"
    ws["K4"] = f'=SUMPRODUCT(--(I{start_row}:I{end_row}<>""),--(J{start_row}:J{end_row}<>0))'
    for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="0F172A")
    for addr, color in (("C4", "2563EB"), ("G4", "16A34A"), ("K4", "DC2626")):
        ws[addr].font = Font(bold=True, color=color, size=16)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_idx = start_row
    for b in branches:
        branch_label = f"{b.name} ({b.code})" if b.code else b.name
        for p in products:
            qty = stock_by_key.get((int(b.id), int(p.id)), 0)
            values = [
                b.id,
                p.id,
                branch_label,
                p.name,
                p.sku or "",
                p.barcode or "",
                "نشط" if p.active else "غير نشط",
                qty,
                None,
                f'=IF(I{row_idx}="","",I{row_idx}-H{row_idx})',
                "",
                int(p.quantity or 0),
                int(p.buy_price or 0),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="right" if col_idx in (4, 11) else "center",
                    vertical="center",
                )
                if col_idx == 9:
                    cell.fill = input_fill
                    cell.font = Font(bold=True)
            row_idx += 1
    if data_rows_count == 0:
        ws.cell(row=start_row, column=1, value="لا توجد منتجات للجرد")

    for row_no in range(start_row, end_row + 1):
        for col_idx in (8, 9, 10, 12, 13):
            ws.cell(row=row_no, column=col_idx).number_format = "#,##0"
    for idx, width in enumerate([10, 10, 22, 34, 14, 18, 12, 16, 16, 12, 24, 16, 14], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.auto_filter.ref = f"A6:M{end_row}"

    summary_headers = ["رقم الفرع", "الفرع", "أسطر الجرد", "المدخل", "أسطر الفرق", "الزيادة", "النقص", "صافي الفرق"]
    summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_headers))
    summary["A1"] = "ملخص الفروقات حسب الفرع"
    summary["A1"].fill = title_fill
    summary["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    summary["A1"].alignment = Alignment(horizontal="center")
    for col_idx, header in enumerate(summary_headers, 1):
        cell = summary.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for idx, b in enumerate(branches, 4):
        summary.cell(row=idx, column=1, value=b.id)
        summary.cell(row=idx, column=2, value=b.name)
        summary.cell(row=idx, column=3, value=f'=COUNTIF(\'الجرد\'!$A${start_row}:$A${end_row},A{idx})')
        summary.cell(row=idx, column=4, value=f'=COUNTIFS(\'الجرد\'!$A${start_row}:$A${end_row},A{idx},\'الجرد\'!$I${start_row}:$I${end_row},"<>")')
        summary.cell(row=idx, column=5, value=f'=SUMPRODUCT(--(\'الجرد\'!$A${start_row}:$A${end_row}=A{idx}),--(\'الجرد\'!$I${start_row}:$I${end_row}<>""),--(\'الجرد\'!$J${start_row}:$J${end_row}<>0))')
        summary.cell(row=idx, column=6, value=f'=SUMIFS(\'الجرد\'!$J${start_row}:$J${end_row},\'الجرد\'!$A${start_row}:$A${end_row},A{idx},\'الجرد\'!$J${start_row}:$J${end_row},">0")')
        summary.cell(row=idx, column=7, value=f'=ABS(SUMIFS(\'الجرد\'!$J${start_row}:$J${end_row},\'الجرد\'!$A${start_row}:$A${end_row},A{idx},\'الجرد\'!$J${start_row}:$J${end_row},"<0"))')
        summary.cell(row=idx, column=8, value=f"=F{idx}-G{idx}")
    total_row = 4 + len(branches)
    summary.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    summary.cell(row=total_row, column=1, value="الإجمالي")
    for col_idx in range(3, 9):
        col = get_column_letter(col_idx)
        summary.cell(row=total_row, column=col_idx, value=f"=SUM({col}4:{col}{total_row - 1})")
    for row in summary.iter_rows(min_row=3, max_row=total_row, min_col=1, max_col=len(summary_headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in summary[total_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True)
    for idx, width in enumerate([12, 26, 14, 12, 12, 12, 12, 14], 1):
        summary.column_dimensions[get_column_letter(idx)].width = width

    info.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    info["A1"] = "تعليمات استخدام ملف الجرد"
    info["A1"].fill = title_fill
    info["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    info["A1"].alignment = Alignment(horizontal="center")
    instructions = [
        ("1", "لا تعدل الأعمدة الزرقاء أو الرمادية، اكتب فقط في عمود الكمية الفعلية الأصفر."),
        ("2", "الفرق ينحسب تلقائياً: الكمية الفعلية - كمية النظام."),
        ("3", "الفرق الموجب يعني زيادة، والفرق السالب يعني نقص."),
        ("4", "اكتب سبب الفرق أو أي ملاحظة في عمود ملاحظة الجرد."),
        ("5", "بعد إكمال الجرد، أرسل نفس الملف حتى أطلع الفروقات حسب الفرع والمنتج."),
    ]
    for row_no, (no, text) in enumerate(instructions, 3):
        info.cell(row=row_no, column=1, value=no)
        info.cell(row=row_no, column=2, value=text)
    for row in info.iter_rows(min_row=3, max_row=3 + len(instructions) - 1, min_col=1, max_col=2):
        for cell in row:
            cell.fill = light_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    info.column_dimensions["A"].width = 10
    info.column_dimensions["B"].width = 80

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, scope



# ======================================
# Inventory Page
# ======================================
@inventory_bp.route("/audit-template.xlsx", methods=["GET"])
def download_inventory_audit_template():
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403

    ensure_product_schema()
    ensure_branch_schema()
    from models.branch import Branch

    branch_id_raw = (request.args.get("branch_id") or "all").strip()
    all_branches = branch_id_raw in ("", "all")
    branch = None
    if not all_branches:
        if not branch_id_raw.isdigit():
            return jsonify({"ok": False, "error": "يرجى اختيار فرع صحيح."}), 400
        branch = Branch.query.filter_by(id=int(branch_id_raw), is_active=True).first()
        if not branch:
            return jsonify({"ok": False, "error": "الفرع غير موجود أو غير نشط."}), 404

    output, scope = _build_inventory_audit_workbook(branch, all_branches=all_branches)
    suffix = "all_branches" if all_branches else _safe_xlsx_filename(scope)
    filename = f"inventory_audit_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@inventory_bp.route("/audit-ai-upload", methods=["POST"])
def upload_inventory_audit_to_ai():
    if not check_permission("can_manage_inventory"):
        return jsonify({"success": False, "error": "غير مصرح"}), 403

    try:
        from utils.ai_assistant_service import (
            build_inventory_reconcile_plan,
            ensure_ai_assistant_schema,
            save_uploaded_file,
        )

        ensure_ai_assistant_schema()
        uploaded = save_uploaded_file(
            request.files.get("file"),
            employee_id=session.get("user_id"),
            session_id=None,
        )
        if uploaded.status != "parsed":
            db.session.commit()
            return jsonify(
                {
                    "success": False,
                    "error": uploaded.error_message or "تعذر قراءة ملف الجرد",
                    "file": {
                        "id": uploaded.id,
                        "original_name": uploaded.original_name,
                        "status": uploaded.status,
                        "preview": uploaded.get_preview(),
                    },
                }
            ), 400

        plan, meta = build_inventory_reconcile_plan(
            [uploaded.id],
            employee_id=session.get("user_id"),
            session_id=None,
        )
        db.session.commit()
        return jsonify(
            {
                "success": True,
                "file": {
                    "id": uploaded.id,
                    "original_name": uploaded.original_name,
                    "status": uploaded.status,
                    "preview": uploaded.get_preview(),
                },
                "plan": plan.to_dict() if plan else None,
                "meta": meta,
                "assistant_url": "/assistant/chat",
            }
        )
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400


@inventory_bp.route("/", methods=["GET", "POST"])
def inventory():
    # فحص الصلاحية
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403

    ensure_product_schema()

    # ==========================
    # ADD PRODUCT (NO SUPPLIER)
    # ==========================
    if request.method == "POST" and request.form.get("form_type") == "add_product":
        opening_stock = int(request.form.get("opening_stock", 0) or 0)
        buy_price = int(request.form.get("buy_price", 0) or 0)
        barcode = request.form.get("barcode", "").strip() or None
        opening_stock_value = opening_stock * buy_price  # قيمة المخزون الافتتاحي
        
        p = Product(
            name=request.form["name"],
            barcode=barcode,
            buy_price=buy_price,
            sale_price=int(request.form["sale_price"]),
            shipping_cost=0,  # إلغاء الحقل
            marketing_cost=0,  # إلغاء الحقل
            opening_stock=opening_stock,
            quantity=opening_stock,  # الكمية تساوي المخزون الافتتاحي بالبداية
            active=True
        )
        db.session.add(p)
        db.session.flush()
        branch_id = _inventory_branch_id()
        if branch_id:
            set_opening_branch_stock(branch_id, p.id, opening_stock)
        
        # ==========================
        # تصحيح محاسبي: المخزون الافتtاحي لا يُسجل كحركة مالية
        # السبب المحاسبي:
        # - المخزون الافتتاحي يُعتبر قيمة مخزون فقط (Asset)
        # - لا يؤثر على الرصيد المالي (Cash/Balance)
        # - لا يظهر في صفحة الحسابات أو سجل الحركات المالية
        # - لا يُعتبر إيداع مالي أو حركة حسابات
        # ==========================
        # تم إزالة إنشاء AccountTransaction للمخزون الافتتاحي
        # المخزون الافتتاحي يُسجل في Product.opening_stock و Product.quantity فقط
        
        db.session.commit()
        return redirect(url_for("inventory.inventory"))

    # ==========================
    # PURCHASE - تم نقله إلى /purchases
    # تم نقل منطق الشراء إلى صفحة purchases منفصلة
    # ==========================

    # ==============================
    # DISPLAY DATA
    # ==============================
    products = Product.query.all()
    view_all = getattr(g, "view_all_branches", False)
    branch_id = _inventory_branch_id()
    stock_map = branch_stock_map(None if view_all else branch_id)
    branch_scope_label = "كل الفروع" if view_all else (g.branch.name if getattr(g, "branch", None) else "—")

    # إحصائيات محسّنة (معالجة آمنة لـ None)
    def _val(v, default=0):
        return default if v is None else v

    def _qty(p):
        return _product_display_qty(p, stock_map, view_all)

    total_purchase = sum(_val(p.buy_price) * _qty(p) for p in products)
    total_sale = sum(_val(p.sale_price) * _qty(p) for p in products)
    total_profit = total_sale - total_purchase

    current_inventory_value = sum(_val(p.buy_price) * _qty(p) for p in products)

    expected_profit_from_stock = sum(
        (_val(p.sale_price) - _val(p.buy_price)) * _qty(p)
        for p in products
    )
    
    low_stock_products = [p for p in products if _qty(p) <= _val(p.low_stock_threshold, 5)]
    low_stock_count = len(low_stock_products)
    
    # المنتجات غير المباعة (quantity > 0 لكن لم تُباع)
    products_with_sales = set()
    from models.order_item import OrderItem
    sold_products = db.session.query(OrderItem.product_id).distinct().all()
    products_with_sales = {p[0] for p in sold_products}
    unsold_products = [p for p in products if _qty(p) > 0 and p.id not in products_with_sales]
    unsold_count = len(unsold_products)
    
    # المنتجات النشطة وغير النشطة
    active_count = sum(1 for p in products if p.active)
    inactive_count = len(products) - active_count
    company_branches = _company_branches_for_form()
    product_branch_map = _product_branch_map(products, company_branches)
    product_branch_stock_map = _product_branch_stock_map(products)

    return render_template(
        "inventory.html",
        products=products,
        total_purchase=total_purchase,
        total_sale=total_sale,
        total_profit=total_profit,
        current_inventory_value=current_inventory_value,
        expected_profit_from_stock=expected_profit_from_stock,
        low_stock_count=low_stock_count,
        low_stock_products=low_stock_products,
        unsold_count=unsold_count,
        unsold_products=unsold_products[:10],  # أول 10 منتجات غير مباعة
        active_count=active_count,
        inactive_count=inactive_count,
        branch_scope_label=branch_scope_label,
        branch_stock_map=stock_map,
        view_all_branches=view_all,
        company_branches=company_branches,
        product_branch_map=product_branch_map,
        product_branch_stock_map=product_branch_stock_map,
    )


# ======================================
# Add Product — صفحة مخصصة (نموذج متقدم)
# ======================================
@inventory_bp.route("/add", methods=["GET", "POST"])
@inventory_bp.route("/add/", methods=["GET", "POST"])
def add_product_page():
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403

    ensure_product_schema()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        if not name:
            ctx = _inventory_add_summary()
            ctx["error"] = "يرجى إدخال اسم المنتج."
            ctx["edit_product"] = None
            ctx["product_meta"] = {}
            ctx["product_specs_items"] = []
            by_prov, def_fee = delivery_fees_from_form(request.form)
            ctx.update({"delivery_by_province": by_prov, "delivery_default_fee": def_fee})
            eid = (request.form.get("edit_product_id") or "").strip()
            if eid.isdigit():
                ep = Product.query.get(int(eid))
                if ep:
                    ctx["edit_product"] = ep
                    ctx["product_meta"] = _load_product_meta(ep)
                    ctx["product_specs_items"] = _extract_specs_items(ctx["product_meta"])
            return render_template("inventory_add_product.html", **ctx), 400

        opening_stock = int(request.form.get("opening_stock", 0) or 0)
        has_colors_flag = bool(request.form.get("has_colors"))
        if has_colors_flag:
            opening_stock = sum(qty for _, qty in _color_rows_from_form(request.form))
        buy_price = int(request.form.get("buy_price", 0) or 0)
        sale_price = int(request.form.get("sale_price", 0) or 0)
        barcode = (request.form.get("barcode") or "").strip() or None
        sku = (request.form.get("sku") or "").strip() or None
        not_for_sale_flag = bool(request.form.get("not_for_sale"))
        low_stock_threshold = int(request.form.get("low_stock_threshold", 5) or 5)
        description = (request.form.get("description") or "").strip() or None
        external_image_url = (request.form.get("external_image_url") or "").strip() or None
        skin_type = (request.form.get("skin_type") or "").strip() or None
        usage_type = (request.form.get("usage_type") or "").strip() or None
        requires_patch_test = bool(request.form.get("requires_patch_test"))
        expiry_date = _parse_optional_date(request.form.get("expiry_date"))
        opened_date = _parse_optional_date(request.form.get("opened_date"))
        batch_number = (request.form.get("batch_number") or "").strip() or None

        meta = _meta_from_inventory_add_form(request.form)

        edit_raw = (request.form.get("edit_product_id") or "").strip()
        edit_id = int(edit_raw) if edit_raw.isdigit() else None

        image_url = None
        file = request.files.get("product_image")
        if file and file.filename:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
                upload_folder = os.path.join(current_app.root_path, "static", "uploads", "products")
                os.makedirs(upload_folder, exist_ok=True)
                raw = secure_filename(file.filename)
                safe = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{raw}"
                path = os.path.join(upload_folder, safe)
                file.save(path)
                image_url = f"/static/uploads/products/{safe}"
        if not image_url and external_image_url:
            image_url = external_image_url

        if edit_id:
            p = Product.query.get(edit_id)
            if not p:
                ctx = _inventory_add_summary()
                ctx["error"] = "المنتج غير موجود للتعديل."
                ctx["edit_product"] = None
                ctx["product_meta"] = {}
                ctx["product_specs_items"] = []
                return render_template("inventory_add_product.html", **ctx), 404

            before_product = snapshot_attrs(p, *PRODUCT_SNAPSHOT_FIELDS)
            old_buy_price = p.buy_price
            old_name = p.name

            p.name = name
            p.sku = sku
            p.barcode = barcode
            p.buy_price = buy_price
            p.sale_price = sale_price
            p.low_stock_threshold = max(0, low_stock_threshold)
            p.description = description
            p.skin_type = skin_type
            p.usage_type = usage_type
            p.requires_patch_test = requires_patch_test
            p.expiry_date = expiry_date
            p.opened_date = opened_date
            p.batch_number = batch_number
            p.shipping_cost = max(0, int(meta.get("delivery_default_fee") or 0))
            p.marketing_cost = 0
            p.active = not not_for_sale_flag

            if image_url:
                p.image_url = image_url
            elif external_image_url:
                p.image_url = external_image_url

            if "current_stock" in request.form and not has_colors_flag:
                new_qty = max(0, int(request.form.get("current_stock") or 0))
                stock_branch_id = _branch_id_for_product_stock(request.form, meta)
                if stock_branch_id:
                    set_branch_stock(stock_branch_id, p.id, new_qty)
                else:
                    p.quantity = new_qty

            p.meta_json = json.dumps(meta, ensure_ascii=False) if meta else None
            db.session.flush()
            color_total = _apply_product_colors_from_form(p, request.form)
            if has_colors_flag:
                p.quantity = color_total
                stock_branch_id = _branch_id_for_product_stock(request.form, meta)
                if stock_branch_id:
                    set_branch_stock(stock_branch_id, p.id, color_total)
            if name != old_name:
                sync_product_name_to_order_items(p.id, name)
            db.session.commit()
            try:
                log_mutation(
                    "update",
                    "inventory",
                    "product",
                    p.id,
                    before_product,
                    snapshot_attrs(p, *PRODUCT_SNAPSHOT_FIELDS),
                    f"تعديل منتج: {p.name}",
                )
            except Exception:
                pass

            action = (request.form.get("submit_action") or "save").strip()
            if action == "add_another":
                return redirect(url_for("inventory.add_product_page"))
            return redirect(url_for("inventory.inventory"))

        p = Product(
            name=name,
            sku=sku,
            barcode=barcode,
            buy_price=buy_price,
            sale_price=sale_price,
            shipping_cost=max(0, int(meta.get("delivery_default_fee") or 0)),
            marketing_cost=0,
            opening_stock=opening_stock,
            quantity=opening_stock,
            active=not not_for_sale_flag,
            low_stock_threshold=max(0, low_stock_threshold),
            description=description,
            image_url=image_url,
            skin_type=skin_type,
            usage_type=usage_type,
            requires_patch_test=requires_patch_test,
            expiry_date=expiry_date,
            opened_date=opened_date,
            batch_number=batch_number,
            meta_json=json.dumps(meta, ensure_ascii=False) if meta else None,
        )
        db.session.add(p)
        db.session.flush()
        stock_branch_id = _branch_id_for_product_stock(request.form, meta)
        if has_colors_flag:
            color_total = _apply_product_colors_from_form(p, request.form)
            p.opening_stock = color_total
            p.quantity = color_total
            if stock_branch_id:
                set_opening_branch_stock(stock_branch_id, p.id, color_total)
        elif stock_branch_id:
            set_opening_branch_stock(stock_branch_id, p.id, opening_stock)
        db.session.commit()
        try:
            log_activity(
                "create",
                "inventory",
                f"إضافة منتج: {p.name}",
                entity_type="product",
                entity_id=p.id,
                payload={"product": snapshot_attrs(p, *PRODUCT_SNAPSHOT_FIELDS)},
            )
        except Exception:
            pass

        action = (request.form.get("submit_action") or "save").strip()
        if action == "add_another":
            return redirect(url_for("inventory.add_product_page"))
        # save | opening | group_prices → العودة لقائمة المخزون
        return redirect(url_for("inventory.inventory"))

    ctx = _inventory_add_summary()
    ctx["edit_product"] = None
    ctx["edit_current_stock"] = 0
    ctx["product_meta"] = {}
    ctx["product_specs_items"] = []
    ctx["product_color_rows"] = []
    ctx.update(_delivery_fees_context())
    edit_arg = request.args.get("edit", type=int)
    if edit_arg:
        ep = Product.query.get(edit_arg)
        if ep:
            ctx["edit_product"] = ep
            ctx["product_meta"] = _load_product_meta(ep)
            ctx["edit_current_stock"] = _edit_current_stock_for_product(ep, ctx["product_meta"])
            ctx["product_specs_items"] = _extract_specs_items(ctx["product_meta"])
            from utils.product_color_service import get_product_colors

            ctx["product_color_rows"] = get_product_colors(ep.id)
            ctx.update(_delivery_fees_context(ep, ctx["product_meta"]))
        else:
            ctx["error"] = "المنتج غير موجود."
    return render_template("inventory_add_product.html", **ctx)


# ======================================
# Add Supplier
# ======================================
@inventory_bp.route("/audit", methods=["GET"])
def audit():
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403

    ensure_product_schema()
    products = Product.query.filter_by(active=True).all()
    return render_template("inventory_audit.html", products=products)
@inventory_bp.route("/save-audit", methods=["POST"])
def save_audit():
    if not check_permission("can_manage_inventory"):
        return jsonify({"success": False, "error": "Unauthorized"}), 403
        
    data = request.get_json()
    items = data.get("items", [])
    
    try:
        branch_id = _inventory_branch_id()
        for item in items:
            product = Product.query.get(item["id"])
            if product:
                expected_qty = get_branch_stock(branch_id, product.id) if branch_id else product.quantity
                actual_qty = item["actual_qty"]
                difference = actual_qty - expected_qty
                
                if difference != 0:
                    if branch_id:
                        set_branch_stock(branch_id, product.id, actual_qty)
                    else:
                        product.quantity = actual_qty
                    
                    # Record the adjustment linearly
                    adjustment_value = difference * product.buy_price
                    if adjustment_value > 0:
                        account_tx = AccountTransaction(
                            type="deposit",
                            amount=adjustment_value,
                            note=f"تسوية جرد غير نقدي بزيادة - {product.name} ({difference:+d} وحدة)"
                        )
                    else:
                        account_tx = AccountTransaction(
                            type="withdraw",
                            amount=abs(adjustment_value),
                            note=f"تسوية جرد غير نقدي بعجز - {product.name} ({difference:+d} وحدة)"
                        )
                    db.session.add(account_tx)
                    
        db.session.commit()
        try:
            log_activity(
                "update",
                "inventory",
                f"حفظ جرد مخزون — {len(items)} منتج",
                payload={"items": items},
            )
        except Exception:
            pass
        return jsonify({"success": True, "message": "تم حفظ تقرير الجرد بنجاح"})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

# ======================================
@inventory_bp.route("/add-supplier", methods=["POST"])
def add_supplier():
    supplier = Supplier(
        name=request.form["name"],
        phone=request.form.get("phone"),
        address=request.form.get("address")
    )
    db.session.add(supplier)
    db.session.commit()
    return redirect(url_for("inventory.inventory"))


# ======================================
# Toggle Product
# ======================================
@inventory_bp.route("/toggle/<int:id>")
def toggle_product(id):
    p = Product.query.get_or_404(id)
    p.active = not p.active
    db.session.commit()
    return redirect(url_for("inventory.inventory"))


# ======================================
# Delete Product
# ======================================
@inventory_bp.route("/delete/<int:id>")
def delete_product(id):
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403

    p = Product.query.get(id)
    if not p:
        flash("المنتج غير موجود أو تم حذفه مسبقاً.", "warning")
        return redirect(url_for("inventory.inventory"))

    try:
        from models.branch import BranchStock, StockTransferLine
        from models.beauty_service_product import BeautyServiceProduct
        from models.maintenance_record import MaintenanceRecord
        from models.purchase import Purchase
        from models.purchase_item import PurchaseItem

        product_id = p.id

        def has_product_link(model):
            try:
                return (
                    db.session.query(model.id)
                    .filter(model.product_id == product_id)
                    .first()
                    is not None
                )
            except SQLAlchemyError:
                db.session.rollback()
                current_app.logger.warning(
                    "Skipping product link check for %s",
                    getattr(model, "__tablename__", model.__name__),
                    exc_info=True,
                )
                return False

        linked_models = (
            OrderItem,
            Purchase,
            PurchaseItem,
            StockTransferLine,
            MaintenanceRecord,
            BeautyServiceProduct,
        )
        has_history = any(has_product_link(model) for model in linked_models)

        if has_history:
            p.active = False
            flash_message = "تم إخفاء المنتج من البيع والمتجر بدلاً من حذفه لأنه مرتبط بسجلات سابقة."
            flash_category = "warning"
        else:
            BranchStock.query.filter_by(product_id=product_id).delete(synchronize_session=False)
            db.session.delete(p)
            flash_message = "تم حذف المنتج بنجاح."
            flash_category = "success"
        try:
            db.session.commit()
            flash(flash_message, flash_category)
        except IntegrityError:
            db.session.rollback()
            p = Product.query.get_or_404(id)
            p.active = False
            db.session.commit()
            flash("تم إخفاء المنتج من البيع والمتجر لأنه مرتبط بسجلات أخرى.", "warning")
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception("delete_product failed")
        flash(f"تعذر حذف المنتج: {exc}", "error")
    return redirect(url_for("inventory.inventory"))


# ======================================
# Edit Product
# ======================================
@inventory_bp.route("/edit/<int:id>", methods=["POST"])
def edit_product(id):
    p = Product.query.get_or_404(id)
    meta = _load_product_meta(p)

    # حفظ القيم القديمة قبل التحديث
    old_name = p.name

    p.name = request.form["name"]
    p.sku = request.form.get("sku", "").strip() or None
    p.barcode = request.form.get("barcode", "").strip() or None
    p.buy_price = int(request.form["buy_price"])
    p.sale_price = int(request.form["sale_price"])
    p.low_stock_threshold = int(request.form.get("low_stock_threshold", 5) or 5)
    p.shipping_cost = 0  # إلغاء الحقل
    p.marketing_cost = 0  # إلغاء الحقل
    p.description = request.form.get("description", "").strip() or None
    p.image_url = request.form.get("image_url", "").strip() or None

    video_url = request.form.get("video_url", "").strip()
    gallery_urls = _split_multiline_values(request.form.get("gallery_urls"))
    specs_items = _specs_items_from_form(request.form)
    store_badge = request.form.get("store_badge", "").strip()

    for key, value in (
        ("video_url", video_url),
        ("store_badge", store_badge),
    ):
        if value:
            meta[key] = value
        else:
            meta.pop(key, None)
    if specs_items:
        meta["specs_items"] = specs_items
    else:
        meta.pop("specs_items", None)
        meta.pop("specs_text", None)
    if gallery_urls:
        meta["gallery"] = gallery_urls
    else:
        meta.pop("gallery", None)
    p.meta_json = json.dumps(meta, ensure_ascii=False) if meta else None
    
    # تحديث المخزون الحالي إذا تم توفيره (لا نعدّل المخزون الافتتاحي)
    if "current_stock" in request.form:
        new_qty = max(0, int(request.form["current_stock"]) if request.form["current_stock"] else 0)
        branch_id = _branch_id_for_product_stock(meta=meta)
        if branch_id:
            set_branch_stock(branch_id, p.id, new_qty)
        else:
            p.quantity = new_qty

    if p.name != old_name:
        sync_product_name_to_order_items(p.id, p.name)
        
    db.session.commit()
    return redirect(url_for("inventory.inventory"))


# ======================================
# Update Opening Stock
# ======================================
@inventory_bp.route("/update-opening-stock/<int:id>", methods=["POST"])
def update_opening_stock(id):
    """تحديث المخزون الافتتاحي"""
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403
    
    product = Product.query.get_or_404(id)
    data = request.get_json() or request.form
    
    try:
        opening_stock = int(data.get("opening_stock", 0))
        product.opening_stock = opening_stock
        db.session.commit()
        return jsonify({"success": True, "message": "تم تحديث المخزون الافتتاحي بنجاح"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 400


# ======================================
# Adjust Stock (تعديل المخزون)
# ======================================
@inventory_bp.route("/adjust-stock/<int:id>", methods=["POST"])
def adjust_stock(id):
    """
    تعديل المخزون يدوياً (مع سبب إلزامي)
    
    السبب المحاسبي:
    - كل حركة مخزون يجب أن تكون مرتبطة بسبب واضح
    - التعديل اليدوي يحتاج سبب لتتبع التغييرات
    - منع أي تغيير مباشر بدون تسجيل حركة
    """
    if not check_permission("can_manage_inventory"):
        return redirect("/pos"), 403
    
    product = Product.query.get_or_404(id)
    
    # الحصول على البيانات (JSON أو Form)
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form.to_dict()
    
    try:
        adjustment = int(data.get("adjustment", 0))  # يمكن أن يكون موجب أو سالب
        reason = data.get("reason", "").strip()
        branch_id = None
        branch_id_raw = str(data.get("branch_id") or "").strip()
        if branch_id_raw.isdigit():
            from models.branch import Branch

            branch = Branch.query.filter_by(id=int(branch_id_raw), is_active=True).first()
            if branch:
                branch_id = branch.id
        if not branch_id:
            branch_id = _inventory_branch_id()
        
        # التحقق من وجود السبب
        if not reason:
            if request.is_json:
                return jsonify({
                    "success": False, 
                    "error": "يجب إدخال سبب التعديل (مثال: فحص جرد، تلف، خطأ في الإدخال)"
                }), 400
            else:
                # Form submission - redirect with error
                from flask import flash
                flash("يجب إدخال سبب التعديل", "error")
                return redirect(url_for("inventory.inventory"))
        
        current_qty = get_branch_stock(branch_id, product.id) if branch_id else (product.quantity or 0)
        if current_qty + adjustment < 0:
            if request.is_json:
                return jsonify({
                    "success": False, 
                    "error": f"المخزون لا يمكن أن يكون سالباً. المخزون الحالي: {current_qty}"
                }), 400
            else:
                from flask import flash
                flash(f"المخزون لا يمكن أن يكون سالباً. المخزون الحالي: {current_qty}", "error")
                return redirect(url_for("inventory.inventory"))
        
        # تطبيق التعديل
        old_quantity = current_qty
        if branch_id:
            adjust_branch_stock(branch_id, product.id, adjustment)
            new_quantity = get_branch_stock(branch_id, product.id)
        else:
            product.quantity += adjustment
            new_quantity = product.quantity
        
        db.session.commit()
        
        # تسجيل الحركة (للعرض فقط - لا نحفظها في قاعدة بيانات)
        # يمكن إضافة سجل حركات في المستقبل إذا لزم الأمر
        
        if request.is_json:
            return jsonify({
                "success": True, 
                "message": f"تم تعديل المخزون بنجاح. السبب: {reason}",
                "old_quantity": old_quantity,
                "new_quantity": new_quantity,
                "adjustment": adjustment
            })
        else:
            from flask import flash
            flash(f"تم تعديل المخزون بنجاح. السبب: {reason}", "success")
            return redirect(url_for("inventory.inventory"))
    except ValueError:
        if request.is_json:
            return jsonify({"success": False, "error": "قيمة التعديل غير صحيحة"}), 400
        else:
            from flask import flash
            flash("قيمة التعديل غير صحيحة", "error")
            return redirect(url_for("inventory.inventory"))
    except Exception as e:
        db.session.rollback()
        if request.is_json:
            return jsonify({"success": False, "error": str(e)}), 400
        else:
            from flask import flash
            flash(f"حدث خطأ: {str(e)}", "error")
            return redirect(url_for("inventory.inventory"))


# ======================================
# Get Product Inventory Movements (API)
# ======================================
@inventory_bp.route("/api/movements/<int:product_id>")
def get_product_movements_api(product_id):
    """API للحصول على حركات مخزون منتج محدد"""
    if not check_permission("can_manage_inventory"):
        return jsonify({"error": "Unauthorized"}), 403
    
    movements = get_product_inventory_movements(product_id)
    summary = get_product_inventory_summary(product_id)
    
    return jsonify({
        "summary": summary,
        "movements": movements[:50]  # آخر 50 حركة
    })


# ======================================
# Product Report
# ======================================
@inventory_bp.route("/report/<int:id>")
def product_report(id):
    product = Product.query.get_or_404(id)

    stats = (
        db.session.query(
            Invoice.status,
            func.count(Invoice.id),
            func.sum(OrderItem.total)
        )
        .join(OrderItem, OrderItem.invoice_id == Invoice.id)
        .filter(OrderItem.product_id == id)
        .group_by(Invoice.status)
        .all()
    )

    report = {
        s[0]: {
            "count": s[1],
            "total": s[2] or 0
        }
        for s in stats
    }

    return render_template(
        "inventory_report.html",
        product=product,
        report=report
    )
