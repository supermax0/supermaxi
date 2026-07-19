"""Verified multi-source knowledge and workbook import for Finora Sales AI."""
from __future__ import annotations

import hashlib
import json
import re
from datetime import datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import or_

from extensions import db
from models.product import Product

from .learning import extract_keywords, normalize_text
from .models import AISalesKnowledgeEntry, AISalesLearningImport, AISalesProductProfile


PRODUCT_SHEET_NAMES = {"المنتجات", "products", "product knowledge", "معرفه المنتجات"}
PROBLEM_SHEET_NAMES = {"المشاكل والحلول", "problems", "issues", "problem solutions", "مشاكل وحلول"}


def _items(value, *, limit: int = 50) -> list[str]:
    if isinstance(value, (list, tuple)):
        source = value
    else:
        source = re.split(r"[\n,،؛;|]+", str(value or ""))
    result: list[str] = []
    for item in source:
        cleaned = str(item or "").strip()
        if cleaned and cleaned not in result:
            result.append(cleaned)
        if len(result) >= limit:
            break
    return result


def _objections(value) -> dict[str, str]:
    result: dict[str, str] = {}
    if isinstance(value, dict):
        source = value.items()
    else:
        source = []
        for line in str(value or "").splitlines():
            parts = re.split(r"\s*(?:=|:|：)\s*", line.strip(), maxsplit=1)
            if len(parts) == 2:
                source.append((parts[0], parts[1]))
    for key, answer in source:
        key_text = str(key or "").strip()
        answer_text = str(answer or "").strip()
        if key_text and answer_text:
            result[key_text] = answer_text
    return result


def _truthy(value, default: bool = True) -> bool:
    if value in (None, ""):
        return default
    return normalize_text(str(value)) not in {"0", "لا", "كلا", "false", "no", "inactive", "متوقف"}


def _signature(product_id: int | None, problem: str) -> str:
    payload = f"problem_solution\n{product_id or 0}\n{normalize_text(problem)}"
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def save_problem_entry(
    data: dict,
    *,
    entry: AISalesKnowledgeEntry | None = None,
    source_type: str = "manual",
    source_name: str = "",
    source_row: int | None = None,
) -> AISalesKnowledgeEntry:
    problem = str(data.get("problem") or "").strip()
    solution = str(data.get("solution") or "").strip()
    if not problem or not solution:
        raise ValueError("المشكلة والحل المعتمد مطلوبان")
    product_id = int(data.get("product_id") or 0) or None
    if product_id and not db.session.get(Product, product_id):
        raise ValueError("المنتج المحدد غير موجود")
    signature = _signature(product_id, problem)
    existing = AISalesKnowledgeEntry.query.filter_by(signature=signature).first()
    if existing and (entry is None or existing.id != entry.id):
        entry = existing
    if entry is None:
        entry = AISalesKnowledgeEntry(signature=signature)
        db.session.add(entry)
    entry.kind = "problem_solution"
    entry.product_id = product_id
    entry.title = str(data.get("title") or problem).strip()[:220] or None
    entry.problem = problem[:2000]
    entry.solution = solution[:4000]
    entry.set_keywords(_items(data.get("keywords"), limit=30) or extract_keywords(problem, limit=20))
    entry.set_diagnostic_questions(_items(data.get("diagnostic_questions"), limit=15))
    entry.escalation_text = str(data.get("escalation") or "").strip()[:2000] or None
    entry.signature = signature
    entry.source_type = source_type
    entry.source_name = str(source_name or "").strip()[:255] or None
    entry.source_row = source_row
    entry.quality_score = 100 if source_type == "manual" else 95
    entry.is_active = _truthy(data.get("is_active"), True)
    return entry


def retrieve_business_knowledge(
    customer_message: str,
    *,
    product_ids: list[int] | None = None,
    limit: int = 4,
) -> list[dict]:
    normalized = normalize_text(customer_message)
    if not normalized:
        return []
    wanted_ids = {int(value) for value in (product_ids or []) if value}
    query = AISalesKnowledgeEntry.query.filter(AISalesKnowledgeEntry.is_active.is_(True))
    if wanted_ids:
        query = query.filter(or_(AISalesKnowledgeEntry.product_id.is_(None), AISalesKnowledgeEntry.product_id.in_(wanted_ids)))
    rows = query.order_by(AISalesKnowledgeEntry.quality_score.desc(), AISalesKnowledgeEntry.updated_at.desc()).limit(400).all()
    wanted_keywords = set(extract_keywords(customer_message, limit=20))
    scored: list[tuple[float, AISalesKnowledgeEntry]] = []
    for row in rows:
        row_keywords = set(row.get_keywords()) | set(extract_keywords(f"{row.title or ''} {row.problem}", limit=24))
        overlap = len(wanted_keywords & row_keywords)
        lexical = overlap / (len(wanted_keywords | row_keywords) or 1)
        similarity = SequenceMatcher(None, normalized, normalize_text(row.problem)).ratio()
        product_boost = 0.18 if row.product_id and row.product_id in wanted_ids else 0
        score = lexical * 0.58 + similarity * 0.24 + product_boost + (row.quality_score / 100) * 0.08
        if overlap or similarity >= 0.48:
            scored.append((score, row))
    scored.sort(key=lambda item: (item[0], item[1].quality_score), reverse=True)
    return [row.to_prompt_dict() for _, row in scored[: max(1, min(int(limit or 4), 6))]]


def build_learning_template() -> BytesIO:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "تعليمات"
    instructions.sheet_view.rightToLeft = True
    instructions.append(["Finora Sales AI - ملف التعلم"])
    instructions.append(["املأ ورقة المنتجات والمشاكل والحلول. لا تضع سعر البيع أو المخزون هنا؛ Finora يقرأهما مباشرة من النظام."])
    instructions.append(["لتحديد المنتج استخدم معرف المنتج أو SKU أو الباركود أو الاسم المطابق الموجود في Finora."])
    instructions.append(["اترك الخلية فارغة إذا لا تريد تغيير الحقل الحالي. افصل القيم المتعددة بسطر جديد أو فاصلة."])

    products = workbook.create_sheet("المنتجات")
    products.sheet_view.rightToLeft = True
    products.append([
        "معرف المنتج", "SKU", "الباركود", "اسم المنتج", "الاسم التسويقي", "أسماء بديلة",
        "المواصفات والوصف", "نقاط البيع", "الاستخدام المناسب", "الضمان", "التوصيل",
        "الألوان", "العرض سم", "الارتفاع سم", "العمق سم",
        "الاعتراضات والردود", "ملاحظات للذكاء", "فعال",
    ])

    problems = workbook.create_sheet("المشاكل والحلول")
    problems.sheet_view.rightToLeft = True
    problems.append([
        "معرف المنتج", "اسم المنتج (اختياري)", "المشكلة", "الكلمات والأعراض",
        "أسئلة التشخيص", "الحل المعتمد", "متى يحول لموظف", "فعال",
    ])
    header_fill = PatternFill("solid", fgColor="2563EB")
    for sheet in (products, problems):
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for index, _ in enumerate(sheet[1], start=1):
            sheet.column_dimensions[get_column_letter(index)].width = 22
    instructions.column_dimensions["A"].width = 105
    instructions["A1"].font = Font(size=16, bold=True, color="2563EB")
    for row in instructions.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


HEADER_ALIASES = {
    "product_id": {"معرف المنتج", "id", "product id", "product_id"},
    "sku": {"sku", "كود المنتج", "الكود"},
    "barcode": {"الباركود", "باركود", "barcode"},
    "product_name": {"اسم المنتج", "اسم المنتج اختياري", "product name", "product"},
    "marketing_name": {"الاسم التسويقي", "marketing name"},
    "aliases": {"اسماء بديله", "اسماء البحث البديله", "aliases"},
    "description": {"المواصفات والوصف", "الوصف", "المواصفات", "description", "specifications"},
    "selling_points": {"نقاط البيع", "selling points"},
    "ideal_for": {"الاستخدام المناسب", "ideal for"},
    "warranty": {"الضمان", "warranty"},
    "delivery": {"التوصيل", "delivery"},
    "colors": {"الالوان", "الألوان", "colors", "available colors"},
    "width_cm": {"العرض سم", "العرض", "width cm", "width"},
    "height_cm": {"الارتفاع سم", "الارتفاع", "height cm", "height"},
    "depth_cm": {"العمق سم", "العمق", "depth cm", "depth"},
    "objections": {"الاعتراضات والردود", "ردود الاعتراضات", "objections"},
    "notes": {"ملاحظات للذكاء", "ملاحظات", "ai notes"},
    "problem": {"المشكله", "problem", "issue"},
    "keywords": {"الكلمات والاعراض", "الكلمات الدلاليه", "keywords", "symptoms"},
    "diagnostic_questions": {"اسئله التشخيص", "diagnostic questions", "questions"},
    "solution": {"الحل المعتمد", "الحل", "solution"},
    "escalation": {"متى يحول لموظف", "التحويل للموظف", "escalation"},
    "is_active": {"فعال", "نشط", "active"},
}


def _header_key(value) -> str:
    normalized = normalize_text(str(value or ""))
    for key, aliases in HEADER_ALIASES.items():
        if normalized in {normalize_text(alias) for alias in aliases}:
            return key
    return normalized.replace(" ", "_")


def _sheet_rows(sheet):
    headers = [_header_key(cell.value) for cell in sheet[1]]
    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {headers[index]: cells[index] for index in range(min(len(headers), len(cells))) if headers[index]}
        if any(value not in (None, "") for value in values.values()):
            yield row_number, values


def _find_sheet(workbook, names: set[str]):
    normalized_names = {normalize_text(name) for name in names}
    return next((sheet for sheet in workbook.worksheets if normalize_text(sheet.title) in normalized_names), None)


def _product_indexes() -> dict[str, dict]:
    indexes = {"id": {}, "sku": {}, "barcode": {}, "name": {}}
    for product in Product.query.all():
        indexes["id"][str(product.id)] = product
        if product.sku:
            indexes["sku"][normalize_text(product.sku)] = product
        if product.barcode:
            indexes["barcode"][normalize_text(product.barcode)] = product
        indexes["name"][normalize_text(product.name)] = product
    return indexes


def _find_product(values: dict, indexes: dict[str, dict]) -> Product | None:
    raw_id = values.get("product_id")
    if raw_id not in (None, ""):
        try:
            product = indexes["id"].get(str(int(float(raw_id))))
            if product:
                return product
        except (TypeError, ValueError):
            pass
    for key in ("sku", "barcode", "product_name"):
        raw = str(values.get(key) or "").strip()
        if raw:
            index_key = "name" if key == "product_name" else key
            product = indexes[index_key].get(normalize_text(raw))
            if product:
                return product
    return None


def import_learning_workbook(content: bytes, filename: str) -> dict:
    safe_name = Path(filename or "finora-learning.xlsx").name[:255]
    if len(content) > 8 * 1024 * 1024:
        raise ValueError("حجم ملف Excel يتجاوز 8 MB")
    digest = hashlib.sha256(content).hexdigest()
    log = AISalesLearningImport(file_name=safe_name, file_hash=digest, status="processing")
    db.session.add(log)
    db.session.flush()
    errors: list[str] = []
    try:
        workbook = load_workbook(BytesIO(content), read_only=False, data_only=True)
        product_sheet = _find_sheet(workbook, PRODUCT_SHEET_NAMES)
        problem_sheet = _find_sheet(workbook, PROBLEM_SHEET_NAMES)
        if product_sheet is None and problem_sheet is None:
            raise ValueError("الملف لا يحتوي ورقة المنتجات أو ورقة المشاكل والحلول")
        indexes = _product_indexes()
        profiles = {row.product_id: row for row in AISalesProductProfile.query.all()}

        if product_sheet is not None:
            for row_number, values in _sheet_rows(product_sheet):
                product = _find_product(values, indexes)
                if not product:
                    errors.append(f"المنتجات - سطر {row_number}: لم يتم العثور على المنتج")
                    log.skipped_rows += 1
                    continue
                profile = profiles.get(product.id)
                if profile is None:
                    profile = AISalesProductProfile(product_id=product.id)
                    profiles[product.id] = profile
                    db.session.add(profile)
                if values.get("marketing_name") not in (None, ""):
                    profile.marketing_name = str(values["marketing_name"]).strip()[:200] or None
                if values.get("aliases") not in (None, ""):
                    profile.aliases_json = json.dumps(_items(values["aliases"]), ensure_ascii=False)
                if values.get("description") not in (None, ""):
                    product.description = str(values["description"]).strip()[:6000] or None
                if values.get("selling_points") not in (None, ""):
                    profile.selling_points_json = json.dumps(_items(values["selling_points"]), ensure_ascii=False)
                if values.get("ideal_for") not in (None, ""):
                    profile.ideal_for_json = json.dumps(_items(values["ideal_for"]), ensure_ascii=False)
                if values.get("warranty") not in (None, ""):
                    profile.warranty_text = str(values["warranty"]).strip()[:220] or None
                if values.get("delivery") not in (None, ""):
                    profile.delivery_text = str(values["delivery"]).strip()[:220] or None
                if values.get("colors") not in (None, ""):
                    profile.colors_json = json.dumps(_items(values["colors"]), ensure_ascii=False)
                for field_name in ("width_cm", "height_cm", "depth_cm"):
                    if values.get(field_name) not in (None, ""):
                        try:
                            measurement = float(str(values[field_name]).replace("،", ".").replace(",", "."))
                            if not 0 < measurement <= 1000:
                                raise ValueError
                            setattr(profile, field_name, round(measurement, 2))
                        except (TypeError, ValueError):
                            errors.append(f"المنتجات - سطر {row_number}: {field_name} يجب أن يكون رقماً بين 0 و1000 سم")
                if values.get("objections") not in (None, ""):
                    profile.objections_json = json.dumps(_objections(values["objections"]), ensure_ascii=False)
                if values.get("notes") not in (None, ""):
                    profile.ai_notes = str(values["notes"]).strip()[:4000] or None
                if values.get("is_active") not in (None, ""):
                    profile.is_active = _truthy(values["is_active"])
                log.product_rows += 1

        if problem_sheet is not None:
            for row_number, values in _sheet_rows(problem_sheet):
                problem = str(values.get("problem") or "").strip()
                solution = str(values.get("solution") or "").strip()
                if not problem and not solution:
                    continue
                if not problem or not solution:
                    errors.append(f"المشاكل والحلول - سطر {row_number}: المشكلة والحل مطلوبان")
                    log.skipped_rows += 1
                    continue
                product_requested = any(values.get(key) not in (None, "") for key in ("product_id", "product_name"))
                product = _find_product(values, indexes) if product_requested else None
                if product_requested and not product:
                    errors.append(f"المشاكل والحلول - سطر {row_number}: المنتج المحدد غير موجود")
                    log.skipped_rows += 1
                    continue
                save_problem_entry(
                    {
                        "product_id": product.id if product else None,
                        "problem": problem,
                        "solution": solution,
                        "keywords": values.get("keywords"),
                        "diagnostic_questions": values.get("diagnostic_questions"),
                        "escalation": values.get("escalation"),
                        "is_active": values.get("is_active"),
                    },
                    source_type="excel",
                    source_name=safe_name,
                    source_row=row_number,
                )
                log.problem_rows += 1

        log.status = "completed_with_errors" if errors else "completed"
        log.error_count = len(errors)
        log.set_errors(errors[:100])
        log.completed_at = datetime.utcnow()
        db.session.commit()
        return log.to_dict()
    except Exception as exc:
        db.session.rollback()
        failed = AISalesLearningImport(file_name=safe_name, file_hash=digest, status="failed", error_count=1, completed_at=datetime.utcnow())
        failed.set_errors([str(exc)])
        db.session.add(failed)
        db.session.commit()
        raise
